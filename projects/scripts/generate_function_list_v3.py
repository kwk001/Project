#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从PRD文档提取功能清单并生成Excel文件 - V3版本
修复一级目录显示问题，添加三段式详细描述和针对性验收标准
"""

import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# 结构化章节（需要特殊处理，子章节是"功能描述/业务流程"等）
STRUCTURED_SECTIONS = {
    '3.2', '3.3', '3.7', '3.8', '3.13', '3.14', '3.15', '3.16',
    '3.18', '3.19', '3.20', '3.21', '3.22', '3.23', '3.24', '3.25', '3.26', '3.27'
}

# 章节到功能点的映射（手动维护关键章节的详细功能）
MODULE_FUNCTIONS = {
    '3.1 基础资料管理': [
        {'name': '航班信息管理', 'desc': '维护航班号、航班日期、起降时间、机型等信息，为货物跟踪提供数据支撑'},
        {'name': '物料/货物管理', 'desc': '统一货物编码标准，支持海关申报和库存管理，关联商品编码、货物属性等'},
        {'name': '客户管理', 'desc': '统一管理客户信息，支持业务关联和信用控制，包括货主、货代等'},
        {'name': '仓库基础设置', 'desc': '配置仓库信息、库区库位、容器类型等仓储基础数据'},
        {'name': '资源管理', 'desc': '管理人力资源、设备资源、月台资源等，支持排班和调度'},
        {'name': '审批管理配置', 'desc': '配置审批流程、审批节点、审批人，支持多级审批'},
        {'name': '单据类型管理', 'desc': '配置业务单据类型、编号规则、打印模板'},
    ],
    '3.2 智能卡口控制': [
        {'name': '车牌自动识别', 'desc': '摄像头自动识别车牌号，识别率≥95%，支持人工补录'},
        {'name': '车辆自动称重', 'desc': '地磅自动采集毛重，与预报重量差异超±5%触发预警'},
        {'name': '海关验核对碰', 'desc': '向金关二期发送验核请求，接收布控/放行/禁止指令'},
        {'name': '自动道闸控制', 'desc': '根据验核结果自动控制道闸开启/关闭，记录通行日志'},
        {'name': '布控车辆拦截', 'desc': '布控车辆自动拦截，下发查验指令，推送至查验区域'},
        {'name': '通行记录查询', 'desc': '查询车辆通行记录，支持按时间/车牌/结果筛选'},
        {'name': '卡口设备监控', 'desc': '监控摄像头/地磅/道闸状态，设备故障自动告警'},
    ],
    '3.3 查验预警处置': [
        {'name': '查验预警提示', 'desc': '货物到港前自动提示查验要求，引导车辆至查验区'},
        {'name': '查验任务分配', 'desc': '自动分配查验任务至查验员，支持人工调整'},
        {'name': '查验作业记录', 'desc': '记录查验过程，拍照留档，录入查验结果'},
        {'name': '查验异常处置', 'desc': '异常货物处理流程，支持退运/销毁/扣留等操作'},
        {'name': '查验结果反馈', 'desc': '查验结果实时反馈海关，触发放行或扣留指令'},
        {'name': '查验直放处理', 'desc': '查验正常货物快速放行，减少二次搬运'},
    ],
    '3.4 货物预报管理': [
        {'name': '货物预报录入', 'desc': '录入货物预报信息，包括航班、件数、重量、品名等'},
        {'name': '安检系统对接', 'desc': '对接机场安检系统，前置申报货物信息，接收安检结果'},
    ],
    '3.5 监管仓管理': [
        {'name': '收货记录管理', 'desc': '记录收货信息，关联预报单，生成入库任务'},
        {'name': '上架单管理', 'desc': '生成上架任务，分配库位，指导上架作业'},
        {'name': '库存实时查询', 'desc': '实时查询库存数量、位置、状态，支持多维度筛选'},
        {'name': '库存盘点管理', 'desc': '定期/不定期盘点，盘点差异处理，盈亏调整'},
        {'name': '库存移位管理', 'desc': '库位间货物移位，移位单管理，位置变更记录'},
        {'name': '发货单管理', 'desc': '创建发货单，分配出库任务，指导拣货作业'},
        {'name': '打板作业管理', 'desc': '航空打板作业指导，板号管理，重量体积校验'},
        {'name': '上机确认管理', 'desc': '货物装机确认，航班配载，装机单生成'},
        {'name': '海关查验协同', 'desc': '查验货物定位，查验结果同步，放行后重新入库'},
    ],
    '3.6 货包机集货跟踪': [
        {'name': '空运集货跟踪', 'desc': '实时监控包机航班动态，关联空运单据信息'},
        {'name': '陆运集货跟踪', 'desc': '公路运输全程可视化跟踪，优化路径，监控节点'},
        {'name': '铁路集货跟踪', 'desc': '跟踪铁路货物在途位置与状态，贴合包机时刻表'},
    ],
    '3.7 订单管理': [
        {'name': '订单创建与审核', 'desc': '客户下单，订单信息审核，审核通过生成运单'},
        {'name': '订单变更管理', 'desc': '支持收货信息、装卸货日期等变更，接口下发变更'},
        {'name': '订单拆分管理', 'desc': '按明细/定量模式拆分主订单，支持取消拆分'},
        {'name': '订单可视化跟踪', 'desc': '订单全流程跟踪，集货时间监控，协同看板展示'},
    ],
    '3.8 运单管理': [
        {'name': '运单创建与变更', 'desc': '根据订单生成运单，支持运单信息变更'},
        {'name': '运单跟踪查询', 'desc': '运单全流程跟踪，在途位置监控，节点状态更新'},
        {'name': '运单自动化配置', 'desc': '自动化规则配置，自动分配承运商/车辆/司机'},
        {'name': '运单拆分与配载', 'desc': '运段拆分，整车/零担配载，运力优化'},
    ],
    '3.9 车辆调度': [
        {'name': '车辆预约管理', 'desc': '司机自助预约到场时间，时段管理，预约码生成'},
        {'name': '排队调度管理', 'desc': '智能排队叫号，优先级排序，实时通知司机'},
    ],
    '3.10 运输执行管理': [
        {'name': '运输任务管理', 'desc': '创建运输任务，分配车辆司机，任务状态跟踪'},
        {'name': '交接单管理', 'desc': '货物交接单据管理，电子签名，责任确认'},
        {'name': '签收管理', 'desc': '货物签收确认，异常签收处理，签收凭证'},
        {'name': '运单图片管理', 'desc': '运输过程拍照上传，货物状态留档'},
        {'name': '回单管理', 'desc': '运输回单收集，回单审核，费用结算依据'},
    ],
    '3.11 月台管理': [
        {'name': '月台分配管理', 'desc': '自动分配最优月台，月台占用监控，调度优化'},
        {'name': '装卸作业管理', 'desc': '装卸任务分配，作业进度跟踪，工作量统计'},
    ],
    '3.12 计费管理': [
        {'name': '计费规则设置', 'desc': '配置计费科目、费率、计费公式，支持多种计费模式'},
        {'name': '费用核算（应收）管理', 'desc': '应收费用计算，费用调整，账单生成'},
        {'name': '应付核算管理', 'desc': '应付费用计算，承运商结算，预付管理'},
        {'name': '报价管理', 'desc': '客户报价管理，报价审批，报价模板'},
    ],
    '3.13 运行中心-综合态势大屏': [
        {'name': '综合态势展示', 'desc': '业务/交通/安防/能耗/安全/消防/资产/设备态势综合展示'},
        {'name': '展示内容配置', 'desc': '配置大屏展示内容，指标选择，布局调整'},
        {'name': '业务规则配置', 'desc': '配置预警阈值，异常规则，自动告警'},
    ],
    '3.14 客服业务管理': [
        {'name': '在线客服咨询', 'desc': '客户在线咨询，即时消息回复，问题解答'},
        {'name': '工单管理', 'desc': '异常问题登记，工单分派，处理跟踪，结果回访'},
        {'name': '知识库管理', 'desc': '常见问题库，自助查询，快速回复模板'},
    ],
    '3.15 登录功能': [
        {'name': '账号密码登录', 'desc': '支持用户名/密码登录，密码加密存储'},
        {'name': '短信验证码登录', 'desc': '手机号+短信验证码登录'},
        {'name': '单点登录集成', 'desc': 'CAS/SSO单点登录，统一身份认证'},
    ],
    '3.16 合同管理': [
        {'name': '合同模板管理', 'desc': '合同模板创建、编辑、版本管理'},
        {'name': '合同签订管理', 'desc': '电子签章，在线签署，合同生效'},
        {'name': '合同变更管理', 'desc': '合同条款变更，变更审批，变更记录'},
    ],
    '3.17 园区管理': [
        {'name': '车辆无感通行', 'desc': '车牌识别自动开闸，白名单管理，快速通行'},
        {'name': '设备台账管理', 'desc': '园区设备登记，设备状态监控，维保记录'},
        {'name': '车位管理与引导', 'desc': '车位占用监控，车位引导，停车缴费'},
        {'name': '巡更巡检管理', 'desc': '巡检路线规划，巡检任务执行，异常上报'},
        {'name': '设备知识库', 'desc': '设备资料管理，维修手册，故障案例'},
    ],
    '3.18 接口对接管理': [
        {'name': '金关二期接口', 'desc': '海关申报、验核、布控、放行数据交换'},
        {'name': '安检系统接口', 'desc': '安检申报、结果接收、异常处理'},
        {'name': '航空公司接口', 'desc': '航班动态、舱单数据、货物状态对接'},
        {'name': '场站系统接口', 'desc': '货物进出港、仓储状态、航班信息同步'},
        {'name': '公路/铁路平台接口', 'desc': '多式联运数据交换，运单跟踪对接'},
    ],
    '3.19 分拣线控制管理': [
        {'name': '分拣任务下发', 'desc': 'WMS向PLC下发分拣任务，货物路由规则'},
        {'name': '扫码识别管理', 'desc': '货物条码扫描，自动识别货物信息'},
        {'name': '自动分拣控制', 'desc': 'PLC控制分拣设备，货物分配至指定格口'},
        {'name': '查验货物分流', 'desc': '布控货物自动分流至查验区'},
        {'name': '打板位资源调度', 'desc': '打板位分配，打板任务调度，资源优化'},
        {'name': 'PLC对接与数据通讯', 'desc': '与PLC设备通信，指令下发，状态采集'},
        {'name': '分拣数据分析', 'desc': '分拣效率统计，设备状态监控，异常告警'},
        {'name': '分拣线仓储管理', 'desc': '管理分拣线侧立体库和缓存区，自动分配库位，执行大小件分区存储策略'},
        {'name': '分拣线全流程跟踪', 'desc': '实时追踪货物从上传送带到打板的全链条状态，记录关键节点位置和状态'},
        {'name': '流程调度管理', 'desc': '配置陆侧空转、空侧直通等业务类型分拣规则，自动路由查验货物'},
    ],
    '3.20 场站系统对接': [
        {'name': '场站系统对接', 'desc': '与机场场站系统数据交换，航班、货物、费用信息同步'},
    ],
    '3.21 航空物流公共信息平台': [
        {'name': '出港服务', 'desc': '货物出港信息查询，在线订舱，运价查询'},
        {'name': '进港服务', 'desc': '货物进港查询，提货预约，费用缴纳'},
        {'name': '全程可视化', 'desc': '货物运输全流程跟踪，节点状态推送'},
        {'name': '平台账号管理', 'desc': '企业/个人账号注册，权限划分，安全认证'},
    ],
    '3.22 移动端功能': [
        {'name': '司机端功能', 'desc': '车辆预约、签到排队、任务接收、运输执行'},
        {'name': '员工端功能', 'desc': '作业任务处理、异常上报、消息通知'},
    ],
    '3.23 统一门户': [
        {'name': '应用集成', 'desc': '各业务系统统一入口，单点登录'},
        {'name': '待办集成', 'desc': '各系统待办任务汇总，统一处理'},
        {'name': '消息集成', 'desc': '系统消息统一展示，消息推送'},
    ],
    '3.24 运维服务管理': [
        {'name': 'IT运维工单管理', 'desc': '系统故障报修，工单分派，处理跟踪，满意度评价'},
        {'name': '设施运维工单管理', 'desc': '设施设备报修，维修调度，验收确认'},
        {'name': '巡检计划管理', 'desc': '定期巡检计划，巡检任务执行，巡检记录'},
    ],
    '3.25 系统管理': [
        {'name': '用户管理', 'desc': '用户增删改查，状态管理，密码策略'},
        {'name': '角色权限管理', 'desc': '角色定义，权限分配，数据权限控制'},
        {'name': '操作日志', 'desc': '操作日志记录，日志查询，日志导出'},
        {'name': '登录日志', 'desc': '登录记录，异常登录告警'},
    ],
    '3.26 跨境电商业务管理': [
        {'name': '9610三单对碰', 'desc': '订单/支付单/物流单三单校验，数据匹配'},
        {'name': '9610清单申报', 'desc': '清单数据生成，清单申报，汇总申报'},
        {'name': '9710订单备案', 'desc': 'B2B订单备案，报关单生成，退税申报'},
        {'name': '9810海外仓备案', 'desc': '海外仓企业备案，备货申报，退货管理'},
        {'name': '跨境电商分拣', 'desc': '跨境电商货物特殊分拣规则，查验分流'},
        {'name': '海关数据对接', 'desc': '跨境电商通关服务平台数据交换'},
    ],
    '3.27 区区联动管理': [
        {'name': '转关申请管理', 'desc': '转关单申请，海关审批，施封管理'},
        {'name': '途中监控管理', 'desc': 'GPS位置跟踪，路线偏离预警，超时告警'},
        {'name': '运抵确认管理', 'desc': '联运区运抵确认，自动解封，入库处理'},
        {'name': '综保区账册管理', 'desc': '保税货物账册管理，入区/出区申报'},
        {'name': '联动数据展示', 'desc': '区区联动数据大屏，监控指标展示'},
    ],
}


def extract_prd_content(file_path):
    """读取PRD文档内容"""
    with open(file_path, 'r', encoding='utf-8') as f:
        return f.read()


def extract_chapters(content):
    """提取章节结构"""
    chapters = []
    pattern = r'^### (3\.\d+(?:\.\d+)?)\s+(.+?)$'
    matches = list(re.finditer(pattern, content, re.MULTILINE))

    for i, match in enumerate(matches):
        chapter_num = match.group(1)
        chapter_title = match.group(2).replace('⭐新增', '').replace('⭐完善', '').replace('（v2.0）', '').replace('⭐新增模块', '').strip()
        start = match.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(content)

        chapters.append({
            'num': chapter_num,
            'title': chapter_title,
            'full_title': f"{chapter_num} {chapter_title}",
            'start': start,
            'end': end,
            'content': content[start:end]
        })

    return chapters


def extract_subsections(chapter_content):
    """提取子章节（#### 级别的内容）"""
    subsections = []
    pattern = r'^#### (\d+\.\d+\.\d+)\s+(.+?)$'
    matches = list(re.finditer(pattern, chapter_content, re.MULTILINE))

    for i, match in enumerate(matches):
        sub_num = match.group(1)
        sub_title = match.group(2).replace('⭐新增', '').replace('⭐完善', '').replace('（v2.0）', '').replace('⭐新增模块', '').replace('⭐补充字段', '').strip()
        start = match.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(chapter_content)

        subsections.append({
            'num': sub_num,
            'title': sub_title,
            'content': chapter_content[start:end]
        })

    return subsections


def get_priority(module_name):
    """根据模块名称判断优先级"""
    if any(x in module_name for x in ['基础资料', '智能卡口', '查验预警', '货物预报', '监管仓', '跨境电商']):
        return 'P0'
    elif any(x in module_name for x in ['新增', '完善']):
        return 'P1'
    else:
        return 'P1'


def get_complexity(func_name, module_name, is_new=False):
    """判断复杂度"""
    if '对接' in func_name or '接口' in func_name or 'PLC' in func_name or '控制' in func_name:
        return '高'
    elif is_new:
        return '中'
    elif any(x in func_name for x in ['管理', '配置', '同步']):
        return '中'
    else:
        return '低'



def generate_three_section_desc(func_name, module_name, simple_desc):
    """生成三段式详细描述
    1. 应用场景
    2. 功能描述
    3. 核心逻辑
    """
    # 根据功能名称推断应用场景
    if '管理' in func_name or '配置' in func_name:
        scenario = f"用户需要进行{func_name}业务操作"
    elif '查询' in func_name or '跟踪' in func_name:
        scenario = f"用户需要查看{func_name.replace('查询', '').replace('跟踪', '')}相关信息"
    elif '对接' in func_name or '接口' in func_name:
        scenario = f"系统需要与外部系统进行{func_name}数据交换"
    elif '控制' in func_name:
        scenario = f"用户需要对{func_name.replace('控制', '')}进行自动化控制"
    elif '识别' in func_name:
        scenario = f"系统需要自动识别{func_name.replace('识别', '').replace('自动', '')}信息"
    elif '分拣' in func_name:
        scenario = f"货物需要通过分拣线进行自动化分拣处理"
    elif '监控' in func_name:
        scenario = f"用户需要实时监控{func_name.replace('监控', '')}状态"
    else:
        scenario = f"用户需要进行{func_name}业务操作"

    # 功能描述使用简单描述
    func_desc = simple_desc

    # 核心逻辑根据功能类型生成
    if '查询' in func_name:
        core_logic = "支持多条件筛选，实时返回查询结果，支持导出"
    elif '管理' in func_name:
        core_logic = "提供增删改查功能，支持批量操作，数据校验"
    elif '对接' in func_name or '接口' in func_name:
        core_logic = "通过标准API接口进行数据交换，支持异步处理"
    elif '控制' in func_name:
        core_logic = "接收指令后自动执行，实时反馈执行状态"
    elif '识别' in func_name:
        core_logic = "自动采集数据，智能识别，准确率≥95%"
    elif '分拣' in func_name:
        core_logic = "根据预设规则自动路由，PLC控制执行"
    elif '跟踪' in func_name:
        core_logic = "实时采集位置信息，全程可视化追踪"
    elif '预警' in func_name or '告警' in func_name:
        core_logic = "阈值监控，异常自动触发，多渠道通知"
    else:
        core_logic = "详见PRD文档详细设计"

    return f"1. 应用场景：{scenario}\n2. 功能描述：{func_desc}\n3. 核心逻辑：{core_logic}"


def generate_function_list(prd_file):
    """生成功能清单数据"""
    content = extract_prd_content(prd_file)
    chapters = extract_chapters(content)

    function_list = []
    func_id = 1

    for ch in chapters:
        # 跳过三级章节（如3.9.5是3.9的子章节）
        if ch['num'].count('.') == 2:
            continue

        module_key = ch['full_title']

        # 检查是否有预定义的功能列表
        if module_key in MODULE_FUNCTIONS:
            for func in MODULE_FUNCTIONS[module_key]:
                # 生成三段式详细描述
                three_section_desc = generate_three_section_desc(
                    func['name'], ch['title'], func['desc']
                )
                
                function_list.append({
                    'ID': f'F{func_id:03d}',
                    '一级目录': ch['title'],
                    '二级目录': '',
                    '页面名称': func['name'] + '页面',
                    '功能点': func['name'],
                    '详细描述': three_section_desc,
                    '优先级': get_priority(ch['title']),
                    '复杂度': get_complexity(func['name'], ch['title']),
                    '依赖': '',
                    '验收标准': ''
                })
                func_id += 1
        else:
            # 提取子章节作为功能点
            subsections = extract_subsections(ch['content'])

            if subsections:
                for sub in subsections:
                    # 跳过结构化章节的非功能子章节
                    if ch['num'] in STRUCTURED_SECTIONS and sub['title'] in ['功能描述', '业务流程', '业务规则', '验收标准', '数据库设计', '接口设计']:
                        continue

                    # 提取功能描述
                    desc_match = re.search(r'\*\*功能描述\*\*[\s\S]*?-\s+\*\*功能价值\*\*：(.+?)(?:\n|$)', sub['content'])
                    detail = desc_match.group(1).strip() if desc_match else ''

                    # 如果没有功能描述段落，提取第一段内容
                    if not detail:
                        para_match = re.search(r'^####[\s\S]*?\n\n(.+?)(?:\n\n|\n###|\Z)', sub['content'], re.DOTALL)
                        if para_match:
                            detail = para_match.group(1).strip()[:200]

                    # 生成三段式详细描述
                    three_section_desc = generate_three_section_desc(
                        sub['title'], ch['title'], detail
                    )
                    
                    function_list.append({
                        'ID': f'F{func_id:03d}',
                        '一级目录': ch['title'],
                        '二级目录': sub['title'],
                        '页面名称': sub['title'],
                        '功能点': sub['title'],
                        '详细描述': three_section_desc,
                        '优先级': get_priority(ch['title']),
                        '复杂度': get_complexity(sub['title'], ch['title']),
                        '依赖': '',
                        '验收标准': ''
                    })
                    func_id += 1
            else:
                # 没有子章节，提取章节本身的功能描述
                desc_match = re.search(r'\*\*功能描述\*\*[\s\S]*?-\s+\*\*功能价值\*\*：(.+?)(?:\n|$)', ch['content'])
                detail = desc_match.group(1).strip() if desc_match else ''

                # 生成三段式详细描述
                three_section_desc = generate_three_section_desc(
                    ch['title'], ch['title'], detail
                )
                
                function_list.append({
                    'ID': f'F{func_id:03d}',
                    '一级目录': ch['title'],
                    '二级目录': '',
                    '页面名称': ch['title'] + '页面',
                    '功能点': ch['title'],
                    '详细描述': three_section_desc,
                    '优先级': get_priority(ch['title']),
                    '复杂度': get_complexity(ch['title'], ch['title']),
                    '依赖': '',
                    '验收标准': ''
                })
                func_id += 1

    return function_list


def merge_cells_for_column(worksheet, column, start_row, end_row):
    """
    合并指定列中相同值的单元格
    只合并非空值（即非'/'的单元格）
    """
    from openpyxl.styles import Alignment

    current_value = None
    merge_start = None

    for row in range(start_row, end_row + 1):
        cell_value = worksheet[f'{column}{row}'].value

        # 如果单元格有值且不是'/'
        if cell_value and cell_value != '/':
            # 如果这是一个新的值，开始新的合并区域
            if cell_value != current_value:
                # 先结束之前的合并区域（如果有）
                if merge_start and current_value and merge_start < row - 1:
                    worksheet.merge_cells(f'{column}{merge_start}:{column}{row - 1}')
                    # 设置合并后的对齐方式
                    worksheet[f'{column}{merge_start}'].alignment = Alignment(
                        vertical='center', horizontal='center', wrap_text=True
                    )

                # 开始新的合并区域
                current_value = cell_value
                merge_start = row
        else:
            # 遇到空值或'/'，结束当前合并区域
            if merge_start and current_value and row - 1 >= merge_start:
                worksheet.merge_cells(f'{column}{merge_start}:{column}{row - 1}')
                worksheet[f'{column}{merge_start}'].alignment = Alignment(
                    vertical='center', horizontal='center', wrap_text=True
                )
                merge_start = None
                current_value = None

    # 处理最后一个合并区域
    if merge_start and current_value and end_row >= merge_start:
        worksheet.merge_cells(f'{column}{merge_start}:{column}{end_row}')
        worksheet[f'{column}{merge_start}'].alignment = Alignment(
            vertical='center', horizontal='center', wrap_text=True
        )


def create_excel(function_list, output_file):
    """创建Excel文件"""
    df = pd.DataFrame(function_list)

    # 创建Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='功能清单')

        # 获取工作表
        workbook = writer.book
        worksheet = writer.sheets['功能清单']

        # 设置表头样式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # 设置列宽
        column_widths = {
            'A': 8,   # ID
            'B': 22,  # 一级目录
            'C': 22,  # 二级目录
            'D': 15,  # 页面名称
            'E': 28,  # 功能点
            'F': 55,  # 详细描述
            'G': 8,   # 优先级
            'H': 8,   # 复杂度
            'I': 12,  # 依赖
            'J': 35   # 验收标准
        }

        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        # 设置数据行样式
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)

        # 冻结首行
        worksheet.freeze_panes = 'A2'

        # 合并一级目录单元格（同一模块的功能点合并显示）
        merge_cells_for_column(worksheet, 'B', 2, worksheet.max_row)

    print(f"功能清单已生成: {output_file}")
    print(f"共 {len(function_list)} 个功能点")


if __name__ == '__main__':
    prd_file = '/Users/apple/Desktop/文件/kaiwu-richang/docs/PRD文档_v2.0.md'
    output_file = '/Users/apple/Desktop/文件/kaiwu-richang/docs/功能清单_v2.0.xlsx'

    function_list = generate_function_list(prd_file)
    create_excel(function_list, output_file)

    # 打印统计
    print("\n功能清单统计：")
    print("=" * 60)

    # 按一级目录统计
    from collections import Counter
    module_counts = Counter([f['一级目录'] for f in function_list])
    for module, count in module_counts.most_common():
        print(f"{module:25} {count:3}个功能点")

    print("=" * 60)

    # 打印前20个功能点预览
    print("\n功能清单预览（前20条）：")
    print("=" * 100)
    for func in function_list[:20]:
        print(f"{func['ID']} | {func['一级目录'][:15]:15} | {func['功能点'][:30]}")
