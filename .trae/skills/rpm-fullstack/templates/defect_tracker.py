#!/usr/bin/env python3
"""缺陷跟踪表模板"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def generate_defect_tracker(filepath: str):
    """生成缺陷跟踪表"""
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "缺陷列表"
    _create_defect_list(ws1)

    ws2 = wb.create_sheet("质量指标")
    _create_quality_metrics(ws2)

    ws3 = wb.create_sheet("质量门禁")
    _create_quality_gate(ws3)

    wb.save(filepath)


def _create_defect_list(ws):
    """创建缺陷列表"""
    ws['A1'] = '缺陷跟踪列表'
    ws.merge_cells('A1:L1')
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ['缺陷编号', '缺陷标题', '严重程度', '优先级', '状态', '发现人',
               '负责人', '发现日期', '关闭日期', '关联需求', '关联任务', '修复版本']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _create_quality_metrics(ws):
    """创建质量指标表"""
    ws['A1'] = '质量指标统计'
    ws.merge_cells('A1:F1')
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')


def _create_quality_gate(ws):
    """创建质量门禁表"""
    ws['A1'] = '质量门禁检查'
    ws.merge_cells('A1:D1')
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')


if __name__ == "__main__":
    generate_defect_tracker("./缺陷跟踪表.xlsx")
