#!/usr/bin/env python3
"""任务跟踪表模板"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def generate_task_tracker(filepath: str):
    """生成任务跟踪表"""
    wb = Workbook()

    # 每日任务跟踪
    ws1 = wb.active
    ws1.title = "每日任务跟踪"
    _create_daily_tracker(ws1)

    # 每周任务汇总
    ws2 = wb.create_sheet("每周任务汇总")
    _create_weekly_summary(ws2)

    # 燃尽图数据
    ws3 = wb.create_sheet("燃尽图数据")
    _create_burndown_data(ws3)

    # 阻塞问题
    ws4 = wb.create_sheet("阻塞问题跟踪")
    _create_blocker_tracker(ws4)

    wb.save(filepath)


def _create_daily_tracker(ws):
    """创建每日任务跟踪"""
    ws['A1'] = '每日任务跟踪表'
    ws.merge_cells('A1:K1')
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ['序号', '任务名称', '负责人', '计划工时', '实际工时', '任务状态', '完成%', '优先级', '阻塞问题', '备注']
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    col_widths = [8, 25, 12, 12, 12, 12, 10, 10, 25, 20]
    for idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width


def _create_weekly_summary(ws):
    """创建每周任务汇总"""
    ws['A1'] = '每周任务汇总表'
    ws.merge_cells('A1:L1')
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')


def _create_burndown_data(ws):
    """创建燃尽图数据"""
    ws['A1'] = '燃尽图数据'
    ws.merge_cells('A1:D1')
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

    headers = ['日期', '计划剩余工时', '实际剩余工时', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')


def _create_blocker_tracker(ws):
    """创建阻塞问题跟踪"""
    ws['A1'] = '阻塞问题跟踪表'
    ws.merge_cells('A1:H1')
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')


if __name__ == "__main__":
    generate_task_tracker("./任务跟踪表.xlsx")
