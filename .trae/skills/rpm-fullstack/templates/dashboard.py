#!/usr/bin/env python3
"""项目仪表盘模板 - 多角色视图"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def generate_dashboard(filepath: str):
    """生成项目仪表盘"""
    wb = Workbook()

    # 管理层视图
    ws1 = wb.active
    ws1.title = "管理层视图"
    _create_executive_view(ws1)

    # PM视图
    ws2 = wb.create_sheet("PM视图")
    _create_pm_view(ws2)

    # 研发视图
    ws3 = wb.create_sheet("研发视图")
    _create_dev_view(ws3)

    # 测试视图
    ws4 = wb.create_sheet("测试视图")
    _create_qa_view(ws4)

    wb.save(filepath)


def _create_executive_view(ws):
    """创建管理层视图"""
    ws['A1'] = '项目执行仪表盘（管理层视图）'
    ws.merge_cells('A1:J1')
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # 项目健康指标
    ws['A3'] = '项目健康度'
    ws['A3'].font = Font(size=12, bold=True)

    metrics = [
        ['整体进度', '65%', '黄色', '延期3天'],
        ['里程碑健康', '3/5正常', '黄色', 'M2延期'],
        ['预算消耗', '60%', '绿色', '正常'],
        ['团队饱和度', '88%', '黄色', '接近上限'],
        ['质量指标', '良好', '绿色', '缺陷率正常'],
    ]

    for row_idx, row_data in enumerate(metrics, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 3:
                if value == '绿色':
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                elif value == '黄色':
                    cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                elif value == '红色':
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    # 风险热力图
    ws['A10'] = '风险热力图'
    ws['A10'].font = Font(size=12, bold=True)


def _create_pm_view(ws):
    """创建PM视图"""
    ws['A1'] = '项目管理仪表盘（PM视图）'
    ws.merge_cells('A1:J1')
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')


def _create_dev_view(ws):
    """创建研发视图"""
    ws['A1'] = '研发任务仪表盘（研发视图）'
    ws.merge_cells('A1:J1')
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')


def _create_qa_view(ws):
    """创建测试视图"""
    ws['A1'] = '质量测试仪表盘（测试视图）'
    ws.merge_cells('A1:J1')
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')


if __name__ == "__main__":
    generate_dashboard("./项目仪表盘.xlsx")
