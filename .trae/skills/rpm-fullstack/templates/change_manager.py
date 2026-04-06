#!/usr/bin/env python3
"""变更管理表模板"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def generate_change_manager(filepath: str):
    """生成变更管理表"""
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "变更请求"
    _create_change_request(ws1)

    ws2 = wb.create_sheet("变更审批")
    _create_change_approval(ws2)

    ws3 = wb.create_sheet("变更影响分析")
    _create_impact_analysis(ws3)

    wb.save(filepath)


def _create_change_request(ws):
    """创建变更请求表"""
    ws['A1'] = '需求变更请求表'
    ws.merge_cells('A1:J1')
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ['变更编号', '变更标题', '变更类型', '提出人', '提出日期',
               '变更原因', '影响范围', '预估工时', '紧急程度', '状态']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _create_change_approval(ws):
    """创建变更审批表"""
    ws['A1'] = '变更审批记录'
    ws.merge_cells('A1:H1')
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')


def _create_impact_analysis(ws):
    """创建变更影响分析表"""
    ws['A1'] = '变更影响分析'
    ws.merge_cells('A1:G1')
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')


if __name__ == "__main__":
    generate_change_manager("./变更管理表.xlsx")
