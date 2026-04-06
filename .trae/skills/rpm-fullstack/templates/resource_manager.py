#!/usr/bin/env python3
"""资源负载表模板"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def generate_resource_manager(filepath: str):
    """生成资源负载表"""
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "团队负载"
    _create_team_load(ws1)

    ws2 = wb.create_sheet("产能分析")
    _create_capacity_analysis(ws2)

    ws3 = wb.create_sheet("资源调配建议")
    _create_resource_recommendation(ws3)

    wb.save(filepath)


def _create_team_load(ws):
    """创建团队负载表"""
    ws['A1'] = '团队资源负载表'
    ws.merge_cells('A1:H1')
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ['成员姓名', '角色', '本周可用', '已分配', '实际投入', '负载率', '饱和度', '瓶颈预警']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _create_capacity_analysis(ws):
    """创建产能分析表"""
    ws['A1'] = '产能分析'
    ws.merge_cells('A1:F1')
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')


def _create_resource_recommendation(ws):
    """创建资源调配建议表"""
    ws['A1'] = '资源调配建议'
    ws.merge_cells('A1:E1')
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')


if __name__ == "__main__":
    generate_resource_manager("./资源负载表.xlsx")
