#!/usr/bin/env python3
"""风险管理表模板"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def generate_risk_manager(filepath: str):
    """生成风险管理表"""
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "风险登记"
    _create_risk_register(ws1)

    ws2 = wb.create_sheet("问题跟踪")
    _create_issue_tracker(ws2)

    wb.save(filepath)


def _create_risk_register(ws):
    """创建风险登记表"""
    ws['A1'] = '项目风险登记表'
    ws.merge_cells('A1:M1')
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ['风险编号', '风险描述', '类别', '等级', '概率', '影响',
               '风险分', '缓解措施', '应急计划', '责任人', '状态', '关联任务', '创建日期']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _create_issue_tracker(ws):
    """创建问题跟踪表"""
    ws['A1'] = '问题跟踪表'
    ws.merge_cells('A1:J1')
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')


if __name__ == "__main__":
    generate_risk_manager("./风险管理表.xlsx")
