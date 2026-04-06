#!/usr/bin/env python3
"""里程碑管理表模板"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def generate_milestone_manager(filepath: str):
    """生成里程碑管理表"""
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "里程碑清单"
    _create_milestone_list(ws1)

    ws2 = wb.create_sheet("里程碑详情")
    _create_milestone_detail(ws2)

    ws3 = wb.create_sheet("风险预警")
    _create_risk_alert(ws3)

    wb.save(filepath)


def _create_milestone_list(ws):
    """创建里程碑清单"""
    ws['A1'] = '项目里程碑管理表'
    ws.merge_cells('A1:K1')
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ['里程碑编号', '里程碑名称', '计划日期', '实际日期', '负责人',
               '验收标准', '当前状态', '完成%', '风险等级', '关联交付物', '备注']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _create_milestone_detail(ws):
    """创建里程碑详情"""
    ws['A1'] = '里程碑详细计划'
    ws.merge_cells('A1:H1')
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')


def _create_risk_alert(ws):
    """创建风险预警"""
    ws['A1'] = '里程碑风险预警看板'
    ws.merge_cells('A1:H1')
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')


if __name__ == "__main__":
    generate_milestone_manager("./里程碑管理表.xlsx")
