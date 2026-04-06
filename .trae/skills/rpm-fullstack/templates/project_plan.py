#!/usr/bin/env python3
"""项目计划表模板 - 包含WBS分解、甘特图、关键路径"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta


def generate_project_plan(filepath: str):
    """生成项目计划表"""
    wb = Workbook()

    # 工作表1: WBS分解
    _create_wbs_sheet(wb)

    # 工作表2: 甘特图
    _create_gantt_sheet(wb)

    # 工作表3: 关键路径
    _create_critical_path_sheet(wb)

    # 工作表4: 项目信息
    _create_project_info_sheet(wb)

    wb.save(filepath)


def _create_wbs_sheet(wb):
    """创建WBS分解工作表"""
    ws = wb.active
    ws.title = "WBS分解"

    # 标题
    ws['A1'] = '项目工作分解结构 (WBS)'
    ws.merge_cells('A1:I1')
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # 表头
    headers = ['WBS编号', '任务名称', '任务描述', '负责人', '计划工期(天)',
               '前置任务', '任务类型', '优先级', '状态']
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col)].width = 15

    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 35

    # WBS数据模板
    wbs_data = [
        ['1', '项目启动', '项目启动阶段所有工作', '', '', '', '阶段', '高', '已完成'],
        ['1.1', '需求分析', '收集和分析业务需求', '产品经理', 5, '', '任务', '高', '已完成'],
        ['1.2', '可行性研究', '技术可行性评估', '架构师', 3, '1.1', '任务', '高', '已完成'],
        ['1.3', '项目立项', '项目立项审批', '项目经理', 2, '1.2', '任务', '高', '进行中'],
        ['2', '设计阶段', '系统详细设计', '', '', '', '阶段', '高', '进行中'],
        ['2.1', '概要设计', '系统架构设计', '架构师', 5, '1.3', '任务', '高', '进行中'],
        ['2.2', '详细设计', '模块详细设计', '技术负责人', 7, '2.1', '任务', '高', '待开始'],
        ['2.3', 'UI/UX设计', '界面原型设计', '设计师', 5, '2.1', '任务', '中', '进行中'],
        ['2.4', '设计评审', '设计方案评审', '项目经理', 2, '2.2', '任务', '高', '待开始'],
        ['3', '开发阶段', '系统开发实现', '', '', '', '阶段', '高', '待开始'],
        ['3.1', '环境搭建', '开发环境配置', 'DevOps', 2, '2.4', '任务', '高', '待开始'],
        ['3.2', '前端开发', '用户界面开发', '前端工程师', 15, '3.1', '任务', '高', '待开始'],
        ['3.3', '后端开发', '服务端开发', '后端工程师', 15, '3.1', '任务', '高', '待开始'],
        ['3.4', '接口联调', '前后端接口联调', '全栈工程师', 5, '3.2', '任务', '高', '待开始'],
        ['3.5', '单元测试', '代码单元测试', '开发工程师', 5, '3.4', '任务', '中', '待开始'],
        ['4', '测试阶段', '系统测试验证', '', '', '', '阶段', '高', '待开始'],
        ['4.1', '集成测试', '系统集成测试', '测试工程师', 5, '3.5', '任务', '高', '待开始'],
        ['4.2', '性能测试', '性能压力测试', '测试工程师', 3, '4.1', '任务', '中', '待开始'],
        ['4.3', 'UAT测试', '用户验收测试', '产品经理', 5, '4.2', '任务', '高', '待开始'],
        ['4.4', 'Bug修复', '缺陷修复', '开发工程师', 5, '4.3', '任务', '高', '待开始'],
        ['5', '上线阶段', '系统上线部署', '', '', '', '阶段', '高', '待开始'],
        ['5.1', '部署准备', '生产环境准备', 'DevOps', 2, '4.4', '任务', '高', '待开始'],
        ['5.2', '数据迁移', '历史数据迁移', 'DBA', 2, '5.1', '任务', '高', '待开始'],
        ['5.3', '系统上线', '正式上线发布', '项目经理', 1, '5.2', '任务', '高', '待开始'],
        ['5.4', '项目验收', '项目验收交付', '项目经理', 3, '5.3', '任务', '高', '待开始'],
    ]

    status_colors = {
        '已完成': 'C6EFCE',
        '进行中': 'FFEB9C',
        '待开始': 'D9D9D9',
        '已阻塞': 'FFC7CE',
        '已延期': 'FFC7CE',
    }

    for row_idx, row_data in enumerate(wbs_data, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

            # 阶段行特殊样式
            if row_data[6] == '阶段':
                cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                cell.font = Font(bold=True)

            # 状态列着色
            if col_idx == 9 and value in status_colors:
                cell.fill = PatternFill(start_color=status_colors[value], end_color=status_colors[value], fill_type='solid')

        ws.row_dimensions[row_idx].height = 22


def _create_gantt_sheet(wb):
    """创建甘特图工作表"""
    ws = wb.create_sheet("甘特图")

    ws['A1'] = '项目甘特图'
    ws.merge_cells('A1:Z1')
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    start_date = datetime(2026, 4, 6)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

    # 表头
    headers = ['任务名称', '负责人', '开始日期', '工期(天)', '结束日期', '完成%']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 日期表头（4周）
    for day in range(28):
        date = start_date + timedelta(days=day)
        col = 7 + day
        cell = ws.cell(row=3, column=col, value=date.strftime('%m/%d'))
        cell.font = Font(bold=True, color='FFFFFF', size=9)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col)].width = 4.5

    # 甘特图数据
    gantt_data = [
        ('需求分析', '产品经理', 0, 5, 1.0),
        ('可行性研究', '架构师', 5, 3, 1.0),
        ('项目立项', '项目经理', 8, 2, 0.5),
        ('概要设计', '架构师', 10, 5, 0.4),
        ('详细设计', '技术负责人', 15, 7, 0.0),
        ('UI/UX设计', '设计师', 15, 5, 0.3),
        ('设计评审', '项目经理', 22, 2, 0.0),
        ('环境搭建', 'DevOps', 24, 2, 0.0),
        ('前端开发', '前端工程师', 26, 15, 0.0),
        ('后端开发', '后端工程师', 26, 15, 0.0),
        ('接口联调', '全栈工程师', 41, 5, 0.0),
        ('单元测试', '开发工程师', 46, 5, 0.0),
        ('集成测试', '测试工程师', 51, 5, 0.0),
        ('性能测试', '测试工程师', 56, 3, 0.0),
        ('UAT测试', '产品经理', 59, 5, 0.0),
        ('Bug修复', '开发工程师', 64, 5, 0.0),
        ('部署准备', 'DevOps', 69, 2, 0.0),
        ('数据迁移', 'DBA', 71, 2, 0.0),
        ('系统上线', '项目经理', 73, 1, 0.0),
        ('项目验收', '项目经理', 74, 3, 0.0),
    ]

    for row_idx, (task, owner, offset, duration, progress) in enumerate(gantt_data, 4):
        ws.cell(row=row_idx, column=1, value=task).border = thin_border
        ws.cell(row=row_idx, column=2, value=owner).border = thin_border
        start = start_date + timedelta(days=offset)
        ws.cell(row=row_idx, column=3, value=start.strftime('%Y-%m-%d')).border = thin_border
        ws.cell(row=row_idx, column=4, value=duration).border = thin_border
        end = start + timedelta(days=duration)
        ws.cell(row=row_idx, column=5, value=end.strftime('%Y-%m-%d')).border = thin_border
        ws.cell(row=row_idx, column=6, value=progress).border = thin_border
        ws.cell(row=row_idx, column=6).number_format = '0%'

        # 甘特条
        for day in range(duration):
            col = 7 + offset + day
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            if day < duration * progress:
                cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            else:
                cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 13
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 13
    ws.column_dimensions['F'].width = 10


def _create_critical_path_sheet(wb):
    """创建关键路径工作表"""
    ws = wb.create_sheet("关键路径")

    ws['A1'] = '关键路径分析'
    ws.merge_cells('A1:H1')
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws['A2'] = '关键路径说明：项目中最长的活动序列，决定项目最短完成时间。关键路径上的任何延误都会导致项目延期。'
    ws.merge_cells('A2:H2')
    ws['A2'].font = Font(color='C00000')
    ws.row_dimensions[2].height = 35

    headers = ['序号', '关键活动', '工期(天)', '最早开始', '最早结束', '最晚开始', '最晚结束', '总浮动']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

    cp_data = [
        [1, '需求分析 → 可行性研究 → 项目立项', 10, '04-06', '04-15', '04-06', '04-15', '0天'],
        [2, '概要设计 → 详细设计 → 设计评审', 14, '04-16', '04-29', '04-16', '04-29', '0天'],
        [3, '环境搭建 → 前端开发 → 接口联调', 21, '04-30', '05-20', '04-30', '05-20', '0天'],
        [4, '后端开发 → 接口联调 → 单元测试', 21, '04-30', '05-20', '04-30', '05-20', '0天'],
        [5, '集成测试 → 性能测试 → UAT测试', 13, '05-21', '06-02', '05-21', '06-02', '0天'],
        [6, 'Bug修复 → 部署准备 → 数据迁移 → 系统上线', 10, '06-03', '06-12', '06-03', '06-12', '0天'],
    ]

    for row_idx, row_data in enumerate(cp_data, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 10


def _create_project_info_sheet(wb):
    """创建项目信息工作表"""
    ws = wb.create_sheet("项目信息")

    ws['A1'] = '项目基本信息'
    ws.merge_cells('A1:C1')
    ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

    info_data = [
        ['项目名称', 'XX系统开发项目（示例）', ''],
        ['项目编号', 'PRJ-2026-001', ''],
        ['项目经理', '张三', ''],
        ['项目周期', '2026-04-06 至 2026-06-15', '共70个工作日'],
        ['项目状态', '进行中', ''],
        ['健康度', '黄色', '存在延期风险'],
        ['', '', ''],
        ['项目成员', '角色', '负载'],
        ['张三', '项目经理', '85%'],
        ['李四', '架构师', '90%'],
        ['王五', '前端工程师', '95%'],
        ['赵六', '后端工程师', '100%'],
        ['钱七', '测试工程师', '75%'],
    ]

    for row_idx, row_data in enumerate(info_data, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 9:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20


if __name__ == "__main__":
    generate_project_plan("./项目计划表.xlsx")
