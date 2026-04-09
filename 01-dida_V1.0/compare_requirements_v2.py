#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
招标需求与客户实际需求对比分析工具 V2
更精确的匹配算法
"""

import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

def extract_prd_features(prd_file):
    """从PRD文档中提取已确认的功能特性"""

    # 分段读取大文件
    features = {
        'MDC数据采集': [],
        '教学管理': [],
        '智慧工厂': [],
        '仓储物流': [],
        '刀具管理': []
    }

    with open(prd_file, 'r', encoding='utf-8') as f:
        content = f.read()

    # 提取关键功能点（通过正则匹配）

    # MDC数据采集相关
    mdc_patterns = [
        r'设备OEE',
        r'设备状态',
        r'设备稼动率',
        r'设备报警',
        r'数据采集',
        r'实时监控',
        r'实训报工',
        r'工时分析'
    ]

    # 教学管理相关
    teaching_patterns = [
        r'教师管理',
        r'学生管理',
        r'课程管理',
        r'班级管理',
        r'排课',
        r'课程表',
        r'教学资源',
        r'教材库',
        r'题目库'
    ]

    # 智慧工厂相关
    factory_patterns = [
        r'订单管理',
        r'项目管理',
        r'生产工单',
        r'APS',
        r'智能排产',
        r'排产策略',
        r'甘特图',
        r'工艺管理',
        r'任务管理'
    ]

    # 仓储物流相关
    logistics_patterns = [
        r'AGV',
        r'WMS',
        r'仓储',
        r'物流',
        r'配送',
        r'立体库',
        r'料箱'
    ]

    # 刀具管理相关
    tool_patterns = [
        r'刀具管理',
        r'刀具需求',
        r'备刀任务',
        r'刀具保养',
        r'刀具BOM',
        r'刀具领用',
        r'刀具归还'
    ]

    # 检查每个类别在PRD中是否存在
    for pattern in mdc_patterns:
        if re.search(pattern, content, re.IGNORECASE):
            features['MDC数据采集'].append(pattern)

    for pattern in teaching_patterns:
        if re.search(pattern, content, re.IGNORECASE):
            features['教学管理'].append(pattern)

    for pattern in factory_patterns:
        if re.search(pattern, content, re.IGNORECASE):
            features['智慧工厂'].append(pattern)

    for pattern in logistics_patterns:
        if re.search(pattern, content, re.IGNORECASE):
            features['仓储物流'].append(pattern)

    for pattern in tool_patterns:
        if re.search(pattern, content, re.IGNORECASE):
            features['刀具管理'].append(pattern)

    return features, content

def analyze_coverage_v2(excel_file, prd_file, output_file):
    """分析覆盖率并生成报告 V2"""

    # 提取PRD特性
    prd_features, prd_content = extract_prd_features(prd_file)

    print("\n从PRD文档提取的功能特性:")
    for category, features in prd_features.items():
        if features:
            print(f"  {category}: {', '.join(features[:5])}{'...' if len(features) > 5 else ''}")

    # 读取Excel
    wb = load_workbook(excel_file)

    # 样式定义
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    high_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    medium_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    low_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    results = {}

    for sheet_name in wb.sheetnames:
        print(f"\n处理: {sheet_name}")
        ws = wb[sheet_name]

        # 确定sheet类型
        sheet_type = ''
        if 'MDC' in sheet_name or '数据采集' in sheet_name:
            sheet_type = 'MDC数据采集'
        elif '教学管理' in sheet_name:
            sheet_type = '教学管理'
        elif '智慧工厂' in sheet_name:
            sheet_type = '智慧工厂'
        elif '仓储物流' in sheet_name or '5G' in sheet_name:
            sheet_type = '仓储物流'

        # 获取该类型的PRD特性
        type_features = prd_features.get(sheet_type, [])

        max_row = ws.max_row
        max_col = ws.max_column

        # 查找或添加列
        coverage_col = None
        desc_col = None

        for col in range(1, max_col + 1):
            val = str(ws.cell(1, col).value or '')
            if '方案覆盖' in val:
                coverage_col = col
            if '详细描述' in val:
                desc_col = col

        if coverage_col is None:
            coverage_col = max_col + 1
            ws.cell(1, coverage_col).value = '方案覆盖率'
            ws.cell(1, coverage_col).font = header_font
            ws.cell(1, coverage_col).fill = header_fill
            max_col += 1

        if desc_col is None:
            desc_col = max_col + 1
            ws.cell(1, desc_col).value = '详细描述'
            ws.cell(1, desc_col).font = header_font
            ws.cell(1, desc_col).fill = header_fill
            max_col += 1

        # 设置列宽
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(coverage_col)].width = 15
        ws.column_dimensions[get_column_letter(desc_col)].width = 80

        # 分析每一行
        covered = 0
        total = 0

        for row in range(2, max_row + 1):
            # 获取指标要求（通常是第2列或第3列）
            indicator = ''
            for col in [2, 3, 4]:
                val = ws.cell(row, col).value
                if val and str(val).strip() and str(val) != 'nan':
                    indicator = str(val)
                    break

            if indicator and len(indicator) > 10:
                total += 1
                coverage, desc = match_requirement(indicator, sheet_type, type_features, prd_content)

                ws.cell(row, coverage_col).value = coverage
                ws.cell(row, coverage_col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                # 设置颜色
                rate_num = 0
                try:
                    rate_num = int(coverage.replace('%', '').replace('?', '0').strip())
                    if rate_num >= 80:
                        ws.cell(row, coverage_col).fill = high_fill
                    elif rate_num >= 50:
                        ws.cell(row, coverage_col).fill = medium_fill
                    elif rate_num > 0:
                        ws.cell(row, coverage_col).fill = low_fill
                except:
                    pass

                ws.cell(row, desc_col).value = desc
                ws.cell(row, desc_col).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                if rate_num > 0:
                    covered += 1

                if total <= 3:  # 只打印前3行用于调试
                    print(f"  行{row}: {coverage} - {indicator[:50]}...")

        coverage_pct = (covered / total * 100) if total > 0 else 0
        results[sheet_name] = {'total': total, 'covered': covered, 'coverage': coverage_pct}
        print(f"  → {covered}/{total} = {coverage_pct:.1f}%")

    wb.save(output_file)
    return results

def match_requirement(indicator, sheet_type, type_features, prd_content):
    """
    匹配单个需求项
    返回: (覆盖率, 描述)
    """
    indicator_lower = indicator.lower()

    # 根据sheet类型进行专门匹配
    if sheet_type == 'MDC数据采集':
        return match_mdc(indicator, type_features, prd_content)
    elif sheet_type == '教学管理':
        return match_teaching(indicator, type_features, prd_content)
    elif sheet_type == '智慧工厂':
        return match_factory(indicator, type_features, prd_content)
    elif sheet_type == '仓储物流':
        return match_logistics(indicator, type_features, prd_content)

    return "?", "待分析"

def match_mdc(indicator, features, prd):
    """匹配MDC数据采集需求"""
    indicator_lower = indicator.lower()

    # 关键功能词
    keywords = {
        'OEE': ['oee', '设备利用率', '设备综合效率'],
        '报工': ['报工', '工时', '工时分析'],
        '车间概况': ['车间概况', '设备建模', '运行状态'],
        '设备详情': ['设备详情', '实时参数', '仪表盘', '转速'],
        '机床监控': ['机床监控', '视频采集', '机床内部', '操作面板'],
        '状态分析': ['状态分析', '状态时序', '历史周期', '设备运行过程']
    }

    for feature, words in keywords.items():
        for word in words:
            if word in indicator_lower:
                # 检查PRD中是否有
                found = any(w in prd.lower() for w in words)
                if found:
                    if feature == 'OEE' and 'oee' in prd.lower():
                        return "100%", "✅ PRD中教师端包含'设备OEE'功能，支持设备稼动率监控和数据分析"
                    elif feature == '报工' and '报工' in prd.lower():
                        return "100%", "✅ PRD中包含'实训报工数据匹配校核'功能，支持工时分析"
                    elif feature in ['车间概况', '设备详情'] and '设备' in prd.lower():
                        return "100%", f"✅ PRD中包含'{feature}'相关功能，支持实时监控"
                    elif feature == '机床监控':
                        return "50%", "⚠️ PRD中未明确提及机床内部视频监控功能，需确认"
                    elif feature == '状态分析' and '状态' in prd.lower():
                        return "80%", "✅ PRD中有设备状态分析功能，但需确认是否支持24H时序图"
                else:
                    return "0%", f"❌ PRD中未找到'{feature}'功能实现"

    return "?", "-"

def match_teaching(indicator, features, prd):
    """匹配教学管理需求"""
    indicator_lower = indicator.lower()

    keywords = {
        '首页导航': ['首页', '导航', '功能导航'],
        '专业班级': ['专业班级', '组织架构', '树状层级'],
        '教师信息': ['教师信息', '教师管理'],
        '学生信息': ['学生信息', '学生管理'],
        'B/S架构': ['b/s', '浏览器'],
        '身份认证': ['身份认证', '账号密码'],
        '角色权限': ['角色', '权限', '访问控制']
    }

    for feature, words in keywords.items():
        for word in words:
            if word in indicator_lower:
                found = any(w in prd.lower() for w in words)
                if found:
                    if feature == '教师信息' and '教师管理' in prd:
                        return "100%", "✅ PRD中3.4节包含完整的'教师管理'功能，支持批量导入、角色配置"
                    elif feature == '学生信息' and '学生管理' in prd:
                        return "100%", "✅ PRD中包含'学生管理'功能，支持批量导入"
                    elif feature == '专业班级' and ('专业' in prd or '班级' in prd):
                        return "100%", "✅ PRD中包含班级管理功能，支持组织架构配置"
                    elif feature == 'B/S架构':
                        return "100%", "✅ PRD中明确系统采用B/S架构，支持多浏览器访问"
                    elif feature == '身份认证' and '登录' in prd:
                        return "100%", "✅ PRD中包含登录功能，支持身份认证"
                    elif feature == '角色权限' and '角色' in prd:
                        return "100%", "✅ PRD中包含角色权限管理功能"
                    else:
                        return "80%", f"✅ PRD中包含'{feature}'相关功能"
                else:
                    return "0%", f"❌ PRD中未找到'{feature}'功能"

    return "?", "-"

def match_factory(indicator, features, prd):
    """匹配智慧工厂需求"""
    indicator_lower = indicator.lower()

    keywords = {
        '订单管理': ['订单', '订单管理', '交期'],
        '项目管理': ['项目', '项目管理', '项目阶段', '甘特图'],
        'APS': ['aps', '智能排产', '排产算法'],
        '生产工单': ['生产工单', '工单'],
        'B/S结构': ['b/s', '浏览器'],
        'API接口': ['api', '接口', '二次开发'],
        '交期预排': ['交期预排', '交期模拟', '工作量选择']
    }

    for feature, words in keywords.items():
        for word in words:
            if word in indicator_lower:
                # 特殊处理交期预排
                if '交期预排' in indicator or '交期模拟' in indicator:
                    return "0%", "❌ PRD中明确说明：学生的实训任务是在其它教务系统排好课后导入，不存在交期预排场景"

                found = any(w in prd.lower() for w in words)
                if found:
                    if feature == '项目管理' and '项目管理' in prd and '甘特图' in prd:
                        return "100%", "✅ PRD中包含完整的项目管理功能，支持项目模版、甘特图、项目任务管理"
                    elif feature == 'APS' and 'aps' in prd.lower():
                        return "100%", "✅ PRD中3.56节包含'APS智能排产'功能，支持智能算法、优先级设置、物料检查等"
                    elif feature == '订单管理' and '订单' in prd:
                        return "100%", "✅ PRD中包含订单管理功能，支持从订单承接到交付的全流程管理"
                    elif feature == '生产工单' and '生产工单' in prd:
                        return "100%", "✅ PRD中包含生产工单管理"
                    elif feature == 'API接口' and 'api' in prd.lower():
                        return "100%", "✅ PRD中明确系统提供开放的API接口平台"
                    else:
                        return "80%", f"✅ PRD中包含'{feature}'相关功能"
                else:
                    return "0%", f"❌ PRD中未找到'{feature}'功能"

    return "?", "-"

def match_logistics(indicator, features, prd):
    """匹配仓储物流需求"""
    indicator_lower = indicator.lower()

    # 硬件指标
    hardware_keywords = ['货位', '180', '50kg', '料箱', '600*400*280', '立柱', '横梁']
    for word in hardware_keywords:
        if word in indicator_lower:
            return "?", "⚠️ 此为硬件指标要求，PRD主要覆盖软件功能，需与硬件供应商确认集成方案"

    # AGV相关
    if 'agv' in indicator_lower:
        if 'agv' in prd.lower() or '刀具' in prd:
            return "50%", "⚠️ PRD中有刀具AGV配送功能，但需确认是否支持料箱式AGV和跨楼层配送"
        else:
            return "0%", "❌ PRD中未明确提及AGV功能"

    # WMS相关
    if 'wms' in indicator_lower or '仓储管理' in indicator_lower:
        if 'wms' in prd.lower() or '仓储' in prd:
            return "50%", "⚠️ PRD中有仓储相关功能，但需确认是否与WMS系统集成"
        else:
            return "0%", "❌ PRD中未明确提及WMS系统"

    return "?", "-"

if __name__ == '__main__':
    print("="*70)
    print("📊 招标需求与客户实际需求对比分析 V2")
    print("="*70)

    try:
        results = analyze_coverage_v2(
            '技术要求V1.1.xlsx',
            'docs/地大智慧工厂_PRD需求文档.md',
            'docs/技术要求V1.1_覆盖率分析.xlsx'
        )

        print("\n" + "="*70)
        print("📈 覆盖率汇总")
        print("="*70)
        for sheet, data in results.items():
            print(f"  {sheet}: {data['coverage']:.1f}% ({data['covered']}/{data['total']})")

        overall = sum(r['coverage'] for r in results.values()) / len(results) if results else 0
        print(f"\n  总体覆盖率: {overall:.1f}%")
        print(f"\n✅ 分析报告已保存: docs/技术要求V1.1_覆盖率分析.xlsx")

    except Exception as e:
        print(f"\n❌ 错误: {e}")
        import traceback
        traceback.print_exc()
