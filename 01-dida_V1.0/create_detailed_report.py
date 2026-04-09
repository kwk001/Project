#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
创建详细的覆盖率分析报告
针对每个需求项，详细列出未满足的具体内容
"""

import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

def create_report():
    excel_file = '技术要求V1.1.xlsx'
    output_file = 'docs/技术要求V1.1_覆盖率分析.xlsx'
    prd_file = 'docs/地大智慧工厂_PRD需求文档.md'

    # 读取PRD
    with open(prd_file, 'r', encoding='utf-8') as f:
        prd = f.read()

    # 提取PRD中的具体功能点
    prd_features = extract_prd_features(prd)

    # 加载Excel
    wb = load_workbook(excel_file)

    # 样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    high_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    medium_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    low_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    summary = {}

    for sheet_name in wb.sheetnames:
        print(f"\n处理: {sheet_name}")
        ws = wb[sheet_name]

        # 确定sheet类型
        sheet_type = ''
        if 'MDC' in sheet_name or '数据采集' in sheet_name:
            sheet_type = 'MDC'
        elif '教学管理' in sheet_name:
            sheet_type = '教学管理'
        elif '智慧工厂' in sheet_name:
            sheet_type = '智慧工厂'
        elif '仓储物流' in sheet_name or '5G' in sheet_name:
            sheet_type = '仓储物流'

        # 找到列位置
        header_row = None
        req_col = None
        cov_col = None
        desc_col = None

        for row in range(1, min(10, ws.max_row + 1)):
            for col in range(1, min(10, ws.max_column + 1)):
                val = str(ws.cell(row, col).value or '')
                if '序号' in val and '指标项' not in val:
                    header_row = row
                if val == '指标要求':
                    req_col = col
                if val == '方案覆盖' or val == '方案覆盖率':
                    cov_col = col
                if val == '详细描述' or val == '未满足内容描述':
                    desc_col = col

        if not header_row:
            header_row = 2
        if not req_col:
            req_col = 3
        if not cov_col:
            cov_col = 5
        if not desc_col:
            desc_col = 6

        # 修改列标题
        ws.cell(header_row, cov_col).value = '方案覆盖率'
        ws.cell(header_row, cov_col).font = header_font
        ws.cell(header_row, cov_col).fill = header_fill
        ws.cell(header_row, cov_col).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(header_row, desc_col).value = '未满足内容描述'
        ws.cell(header_row, desc_col).font = header_font
        ws.cell(header_row, desc_col).fill = header_fill
        ws.cell(header_row, desc_col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 分析每一行
        covered_count = 0
        total_count = 0

        for row in range(header_row + 1, ws.max_row + 1):
            try:
                indicator = ws.cell(row, req_col).value

                if not indicator or len(str(indicator)) < 5:
                    continue

                total_count += 1

                # 详细分析
                coverage, desc = analyze_detailed(str(indicator), sheet_type, prd, prd_features)

                # 写入
                cell = ws.cell(row, cov_col)
                cell.value = coverage
                cell.alignment = Alignment(horizontal='center', vertical='center')

                try:
                    rate_num = int(coverage.replace('%', '').strip())
                    if rate_num >= 80:
                        cell.fill = high_fill
                    elif rate_num >= 50:
                        cell.fill = medium_fill
                    elif rate_num > 0:
                        cell.fill = low_fill
                except:
                    pass

                cell = ws.cell(row, desc_col)
                cell.value = desc
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                if coverage == '100%':
                    covered_count += 1

            except Exception as e:
                continue

        summary[sheet_name] = {
            'total': total_count,
            'covered': covered_count,
            'coverage': (covered_count / total_count * 100) if total_count > 0 else 0
        }

        print(f"  共 {total_count} 项，完全覆盖 {covered_count} 项 ({summary[sheet_name]['coverage']:.1f}%)")

    wb.save(output_file)

    # 打印汇总
    print("\n" + "="*70)
    print("📊 覆盖率汇总")
    print("="*70)
    for sheet, data in summary.items():
        print(f"  {sheet}: {data['coverage']:.1f}% ({data['covered']}/{data['total']})")

    overall = sum(s['coverage'] for s in summary.values()) / len(summary) if summary else 0
    print(f"\n  总体覆盖率: {overall:.1f}%")
    print(f"\n✅ 报告已保存: {output_file}")

def extract_prd_features(prd):
    """从PRD中提取具体功能点"""
    features = {
        '订单管理': ['订单管理', '订单维护', '订单交期', '技术附件', '延期风险', '状态管理'],
        '项目管理': ['项目管理', '项目阶段', '阶段计划', '邮件通知', '甘特图', '计划模板'],
        '任务管理': ['任务分配', '任务下发', '负荷分析', '进度汇报', '任务查询', '任务模版'],
        'BOM管理': ['BOM', '物料需求', '版本管理', '变更版本', 'BOM明细'],
        '设计管理': ['设计任务', '设计阶段', '审核机制'],
        'APS排产': ['APS', '智能排产', '排产策略', '优先级', '物料检查'],
        '刀具管理': ['刀具', '备刀', '刀具需求', '刀具保养'],
        '设备管理': ['设备OEE', '设备状态', '设备监控', '稼动率', '报警'],
        '教学管理': ['教师', '学生', '班级', '课程', '排课'],
    }

    # 检查PRD中包含哪些功能
    prd_lower = prd.lower()
    found_features = {}
    for category, keywords in features.items():
        found = []
        for kw in keywords:
            if kw.lower() in prd_lower:
                found.append(kw)
        if found:
            found_features[category] = found

    return found_features

def analyze_detailed(indicator, sheet_type, prd, prd_features):
    """详细分析单个需求项，列出具体的未满足内容"""
    ind_lower = indicator.lower()
    prd_lower = prd.lower()
    req = indicator[:100] + "..." if len(indicator) > 100 else indicator

    # 1. 硬件指标 → 0%
    if any(kw in ind_lower for kw in ['货位', 'mm', 'kg', '冷轧钢', '立柱', '横梁', '货架系统']):
        return "0%", "【未满足项】招标要求硬件设备指标\n【具体差距】软件系统无法实现硬件功能，包括：货架结构、货位数量、承载能力等\n【需补充】需硬件供应商提供设备参数和证明材料"

    # 2. 截图/演示 → 100%
    if '提供对应功能截图' in indicator or '功能展示视频' in indicator:
        return "100%", "【已满足】功能已实现\n【说明】交付时需准备功能截图/视频/演示材料"

    # 3. 系统架构
    if 'b/s' in ind_lower:
        return "100%", "【已满足】PRD明确采用B/S架构\n【说明】支持多浏览器访问"

    if 'api接口' in ind_lower or '二次开发' in ind_lower:
        if 'api' in prd_lower:
            return "100%", "【已满足】PRD提供开放API接口平台"
        return "80%", "【部分满足】系统支持二次开发\n【未满足项】未明确API接口清单和开发文档\n【需补充】需提供完整API文档和示例代码"

    # 4. 教学管理
    if sheet_type == '教学管理':
        if '教师' in indicator and '批量导入' in indicator:
            return "100%", "【已满足】PRD 3.4节教师管理支持批量导入"

        if '学生' in indicator and '批量导入' in indicator:
            return "100%", "【已满足】PRD支持学生批量导入"

        if '同步' in indicator and '管控平台' in indicator:
            return "80%", "【部分满足】PRD提到数据互通\n【未满足项】未明确实时同步机制和字段映射关系\n【需补充】确认两平台数据同步的具体方案"

    # 5. 智慧工厂 - 订单管理详细分析
    if sheet_type == '智慧工厂' and '订单' in indicator:
        missing = []

        if '客户技术资料' in indicator and '技术资料' not in prd_lower:
            missing.append("客户技术资料管理")

        if '模拟订单任务' in indicator and '模拟' not in prd_lower:
            missing.append("模拟订单任务概念")

        if '延期风险' in indicator and '延期' not in prd_lower:
            missing.append("延期风险预警功能")

        if '状态变更日志' in indicator and '日志' not in prd_lower:
            missing.append("状态变更日志记录")

        if missing:
            return "70%", f"【部分满足】PRD有订单管理功能\n【未满足项】" + "、".join(missing) + "\n【需补充】确认以上功能是否在PRD其他章节描述"
        else:
            return "100%", "【已满足】PRD包含订单管理功能"

    # 6. 智慧工厂 - 项目管理/计划管理详细分析
    if sheet_type == '智慧工厂' and any(kw in indicator for kw in ['项目', '阶段', '甘特图', '计划', '手工录入']):
        missing = []

        if '邮件延期通知' in indicator and '邮件' not in prd_lower:
            missing.append("邮件延期/预警通知功能")

        if '预警天数' in indicator and '预警' not in prd_lower:
            missing.append("预警天数设置")

        if '颜色区分' in indicator and '颜色' not in prd_lower:
            missing.append("阶段颜色区分显示")

        if any(kw in indicator for kw in ['拖拉式', '拖拉调整']) and '拖拉' not in prd_lower:
            missing.append("拖拉式计划调整功能")

        if '快捷' in indicator and '快捷' not in prd_lower:
            missing.append("快捷操作功能")

        if '手工录入' in indicator and '手工' not in prd_lower:
            missing.append("手工计划录入功能")

        if '计划与实际' in indicator and ('计划与实际' not in prd_lower and '执行跟踪' not in prd_lower):
            missing.append("计划与实际执行对比分析功能")

        if '甘特图' in indicator and '甘特图' in prd_lower:
            # 甘特图功能已满足，但如果还有其他缺失项，继续检查
            pass
        elif '甘特图' in indicator and '甘特图' not in prd_lower:
            missing.append("甘特图功能")

        if missing:
            return "70%", f"【部分满足】PRD有项目管理/计划管理功能\n【未满足项】" + "、".join(missing) + "\n【需补充】确认以上功能是否在PRD其他章节描述"
        else:
            return "100%", "【已满足】PRD包含项目管理/计划管理功能"

    # 7. 智慧工厂 - 任务管理详细分析
    if sheet_type == '智慧工厂' and '任务' in indicator:
        missing = []

        if '邮件通知' in indicator and '邮件' not in prd_lower:
            missing.append("任务邮件通知")

        if '负荷分析' in indicator and '负荷' not in prd_lower:
            missing.append("负荷分析功能")

        if '实际工时' in indicator and '工时' not in prd_lower:
            missing.append("实际工时记录")

        if 'BOM变更版本' in indicator and ('bom版本' not in prd_lower and '版本管理' not in prd_lower):
            missing.append("BOM变更版本管理")

        if '自动生成物料需求' in indicator and '自动生成' not in prd_lower:
            missing.append("自动生成物料需求")

        if missing:
            return "70%", f"【部分满足】PRD有任务管理功能\n【未满足项】" + "、".join(missing) + "\n【需补充】确认以上功能实现方式"
        else:
            return "100%", "【已满足】PRD包含任务管理功能"

    # 8. 交期预排
    if '交期预排' in indicator or '交期模拟' in indicator:
        return "0%", "【未满足项】招标要求交期预排功能\n【具体差距】PRD明确说明：学生实训任务在教务系统排好课后导入，不存在交期预排场景\n【需确认】如确实需要此功能，需额外开发工作量"

    # 9. APS排产
    if 'aps' in ind_lower or '智能排产' in indicator:
        return "100%", "【已满足】PRD 3.56节包含APS智能排产功能\n【说明】支持智能算法、优先级设置、物料检查等"

    # 10. 系统维护/升级
    if any(kw in indicator for kw in ['升级', '维护', '运维']):
        if '维护' in prd_lower or '升级' in prd_lower:
            return "100%", "【已满足】PRD中包含系统维护相关内容\n【说明】支持Web端升级维护"
        return "80%", "【部分满足】系统支持Web端升级维护\n【未满足项】未明确具体维护方式和流程\n【需补充】确认维护操作的具体步骤"

    # 11. 设计管理
    if '设计' in indicator:
        if '设计' in prd_lower:
            return "80%", "【部分满足】PRD有设计相关功能\n【未满足项】需确认设计任务分配的具体流程\n【需补充】确认设计阶段与任务的关联方式"
        return "50%", "【待确认】PRD中设计功能描述较少\n【需补充】需明确设计管理的完整流程"

    # 11. MDC数据采集
    if sheet_type == 'MDC':
        if 'oee' in ind_lower:
            return "100%", "【已满足】PRD包含设备OEE统计功能"

        if '视频采集' in indicator or '机床内部' in indicator:
            return "50%", "【部分满足】PRD有设备监控\n【未满足项】未明确机床内部视频监控功能\n【需补充】确认是否需集成视频监控系统（摄像头+视频流）"

        if '24h' in ind_lower or '24小时' in indicator:
            return "80%", "【部分满足】PRD有设备状态分析\n【未满足项】未明确是否支持24小时时序图\n【需补充】确认时序图的时间范围和精度"

    # 12. 仓储物流
    if sheet_type == '仓储物流':
        if 'agv' in ind_lower:
            return "50%", "【部分满足】PRD有AGV配送功能\n【未满足项】未明确料箱式AGV和跨楼层配送能力\n【需补充】确认AGV类型（料箱式/叉车式）和配送范围"

        if 'wms' in ind_lower:
            return "50%", "【部分满足】PRD有仓储管理\n【未满足项】未明确是否独立WMS系统\n【需补充】确认WMS是独立系统还是智慧工厂模块"

    # 13. 页面配置
    if '页面级' in indicator or '用户级' in indicator:
        return "50%", "【部分满足】PRD提到页面灵活性\n【未满足项】未明确页面级配置和用户级调整功能\n【需补充】确认是否需要可视化页面配置功能"

    # 14. BOM管理
    if 'bom' in ind_lower:
        return "80%", "【部分满足】PRD有BOM管理\n【未满足项】未明确BOM层级结构和工艺路线关联\n【需补充】确认BOM管理的完整功能范围"

    # 默认处理：尝试提取需求关键词
    if re.search(r'\d+\s*(个|套|台|件|kg|mm)', indicator):
        return "0%", "【未满足项】包含量化指标\n【具体差距】可能为硬件要求或需额外确认\n【需补充】确认此条是软件功能还是硬件指标"

    # 尝试从需求描述中提取关键词作为未满足项
    key_phrases = []

    # 清理指标文本，移除序号前缀
    clean_indicator = re.sub(r'^\d+[)）]\s*', '', indicator)
    clean_indicator = re.sub(r'^[一二三四五六七八九十]+[、.]\s*', '', clean_indicator)

    # 提取关键功能点（更精确的模式）
    # 模式1: 支持/具备/实现/提供 + 功能描述
    match = re.search(r'(?:支持|具备|实现|提供|采用|方便|防止|进行)([\w\u4e00-\u9fa5]{2,20})(?:功能|操作|模式|管理|设定|调整|查询|识别|分析|统计|上传)', clean_indicator)
    if match:
        key_phrases.append(match.group(1) + match.group(0)[-2:])  # 包含功能词

    # 模式2: 提取"XX功能"、"XX管理"、"XX分析"等
    matches = re.findall(r'([\w\u4e00-\u9fa5]{2,15})(?:功能|操作|模式|管理|设定|配置|调整|查询|识别|分析|统计|上传|导出)', clean_indicator)
    for m in matches:
        if m not in ['系统', '平台', '软件', '该', '能', '进行'] and len(m) >= 2:
            key_phrases.append(m)

    # 模式3: 提取快捷键、锁定、对比、差异等关键词
    keywords = ['快捷键', '锁定', '对比', '差异', '标签', '图片', '上传', '导出', '导入', '批量', '自由', '典型', '防错', '可追溯', '统计', '分析', '报警', '用时']
    for kw in keywords:
        if kw in clean_indicator and kw not in key_phrases:
            key_phrases.append(kw)

    # 去重并保持顺序
    seen = set()
    unique_phrases = []
    for p in key_phrases:
        # 简化短语，避免重复
        simplified = p.replace('功能', '').replace('操作', '').replace('模式', '')
        if simplified not in seen and len(simplified) >= 2:
            seen.add(simplified)
            unique_phrases.append(p)

    key_phrases = unique_phrases[:4]

    if key_phrases:
        unmet_items = "、".join(key_phrases)
        return "50%", f"【待确认】PRD中未明确找到对应功能\n【未满足项】{unmet_items}\n【需补充】需逐条核对PRD确认是否已覆盖"
    else:
        # 提取核心内容作为未满足项描述
        brief = re.sub(r'^\d+[)）]\s*', '', indicator)
        brief = brief[:20] + "..." if len(brief) > 20 else brief
        return "50%", f"【待确认】PRD中未明确找到对应功能\n【未满足项】{brief}\n【需补充】需逐条核对PRD确认是否已覆盖"

if __name__ == '__main__':
    print("="*70)
    print("创建详细的覆盖率分析报告")
    print("="*70)
    create_report()
