#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
清理Excel文件，只保留必要的列
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from copy import copy

def cleanup_excel():
    input_file = 'docs/技术要求V1.1_覆盖率分析.xlsx'
    output_file = 'docs/技术要求V1.1_覆盖率分析.xlsx'

    wb = load_workbook(input_file)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n处理: {sheet_name}")

        # 找到需要保留的列
        keep_cols = {}  # {新列号: (原列号, 标题)}

        for col in range(1, ws.max_column + 1):
            val = str(ws.cell(2, col).value or '')  # 第2行通常是标题行
            if val:
                if '序号' in val:
                    keep_cols[1] = (col, '序号')
                elif '指标项' in val:
                    keep_cols[2] = (col, '指标项')
                elif '指标要求' in val and '覆盖率' not in val:
                    keep_cols[3] = (col, '指标要求')
                elif '方案覆盖率' in val:
                    keep_cols[4] = (col, '方案覆盖率')
                elif '未满足内容描述' in val:
                    keep_cols[5] = (col, '未满足内容描述')

        if len(keep_cols) < 5:
            # 尝试从第1行找
            for col in range(1, ws.max_column + 1):
                val = str(ws.cell(1, col).value or '')
                if '方案覆盖率' in val:
                    keep_cols[4] = (col, '方案覆盖率')
                elif '未满足内容描述' in val:
                    keep_cols[5] = (col, '未满足内容描述')

        print(f"  保留列: {keep_cols}")

        # 创建新数据
        max_row = ws.max_row

        # 收集数据
        data = []
        for row in range(1, max_row + 1):
            row_data = []
            for new_col in range(1, 6):
                if new_col in keep_cols:
                    old_col, _ = keep_cols[new_col]
                    row_data.append(ws.cell(row, old_col).value)
                else:
                    row_data.append(None)
            data.append(row_data)

        # 清空工作表
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None

        # 写回数据
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)

        for row_idx, row_data in enumerate(data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row_idx, col_idx)
                cell.value = value

                # 标题行样式
                if row_idx == 2:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 设置列宽
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(1)].width = 8
        ws.column_dimensions[get_column_letter(2)].width = 20
        ws.column_dimensions[get_column_letter(3)].width = 60
        ws.column_dimensions[get_column_letter(4)].width = 15
        ws.column_dimensions[get_column_letter(5)].width = 80

    wb.save(output_file)
    print(f"\n✅ 已清理: {output_file}")

if __name__ == '__main__':
    cleanup_excel()
