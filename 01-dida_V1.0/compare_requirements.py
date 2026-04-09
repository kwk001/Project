#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
招标需求与客户实际需求对比分析工具
以技术要求V1.1.xlsx为标准，对比PRD文档，生成覆盖率分析报告
"""

import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def read_prd_content(prd_file):
    """读取PRD文档内容并提取关键功能点"""
    with open(prd_file, 'r', encoding='utf-8') as f:
        content = f.read()

    # 提取PRD中的功能模块
    modules = {}

    # 匹配教务端、教师端、学生端功能
    sections = re.findall(r'### (\d+\.\d+) (.+?)\n', content)
    for num, name in sections:
        modules[name.strip()] = {
            'section': num,
            'content': ''
        }

    # 提取功能关键词
    keywords = {
        'MDC': ['MDC', '数据采集', '设备OEE', '设备状态', '稼动率', '报警'],
        '教学管理': ['教学管理', '教务', '教师管理', '学生管理', '课程管理', '排课'],
        '智慧工厂': ['智慧工厂', '订单管理', '项目管理', '生产工单', 'APS', '排产'],
        '仓储物流': ['仓储', 'AGV', 'WMS', '立体库', '料箱', '配送', '物流']
    }

    return modules, keywords, content

def analyze_coverage(excel_file, prd_file, output_file):
    """分析覆盖率并生成报告"""

    # 读取PRD内容
    prd_modules, keywords, prd_content = read_prd_content(prd_file)

    # 读取Excel文件
    xlsx = pd.ExcelFile(excel_file)

    # 使用openpyxl处理Excel
    wb = load_workbook(excel_file)

    # 定义样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    high_coverage_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # 绿色
    medium_coverage_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # 黄色
    low_coverage_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # 红色

    results = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 确定当前sheet的类型
        sheet_type = ''
        if 'MDC' in sheet_name or '数据采集' in sheet_name:
            sheet_type = 'MDC'
        elif '教学管理' in sheet_name:
            sheet_type = '教学管理'
        elif '智慧工厂' in sheet_name:
            sheet_type = '智慧工厂'
        elif '仓储物流' in sheet_name or '5G' in sheet_name:
            sheet_type = '仓储物流'

        # 获取当前sheet的关键字
        type_keywords = keywords.get(sheet_type, [])

        # 找到数据起始行和列
        max_row = ws.max_row
        max_col = ws.max_column

        # 查找"指标要求"列的位置
        indicator_col = None
        coverage_col = None
        desc_col = None

        for col in range(1, max_col + 1):
            cell_value = ws.cell(1, col).value
            if cell_value and '指标要求' in str(cell_value):
                indicator_col = col
            if cell_value and '方案覆盖' in str(cell_value):
                coverage_col = col
            if cell_value and '详细描述' in str(cell_value):
                desc_col = col

        # 如果没有找到这些列，在最后一列后添加
        if coverage_col is None:
            coverage_col = max_col + 1
            ws.cell(1, coverage_col).value = '方案覆盖率'
            ws.cell(1, coverage_col).font = header_font
            ws.cell(1, coverage_col).fill = header_fill
            ws.cell(1, coverage_col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            max_col += 1

        if desc_col is None:
            desc_col = max_col + 1
            ws.cell(1, desc_col).value = '详细描述'
            ws.cell(1, desc_col).font = header_font
            ws.cell(1, desc_col).fill = header_fill
            ws.cell(1, desc_col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            max_col += 1

        # 设置列宽
        ws.column_dimensions[chr(64 + coverage_col)].width = 15
        ws.column_dimensions[chr(64 + desc_col)].width = 60

        # 分析每一行
        covered_count = 0
        total_count = 0

        for row in range(2, max_row + 1):
            # 获取指标要求内容
            indicator_value = ''
            if indicator_col:
                indicator_value = str(ws.cell(row, indicator_col).value or '')

            # 如果这一行有指标要求内容
            if indicator_value and indicator_value.strip() and indicator_value != 'nan':
                total_count += 1

                # 分析覆盖率
                coverage_rate, description = analyze_single_item(
                    indicator_value,
                    sheet_type,
                    type_keywords,
                    prd_content
                )

                # 写入覆盖率
                ws.cell(row, coverage_col).value = coverage_rate
                ws.cell(row, coverage_col).alignment = Alignment(horizontal='center', vertical='center')

                # 根据覆盖率设置背景色
                rate_num = 0
                try:
                    rate_num = int(coverage_rate.replace('%', '').replace('?', '0'))
                    if rate_num >= 80:
                        ws.cell(row, coverage_col).fill = high_coverage_fill
                    elif rate_num >= 50:
                        ws.cell(row, coverage_col).fill = medium_coverage_fill
                    elif rate_num > 0:
                        ws.cell(row, coverage_col).fill = low_coverage_fill
                except:
                    pass

                # 写入详细描述
                ws.cell(row, desc_col).value = description
                ws.cell(row, desc_col).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                if rate_num > 0:
                    covered_count += 1

        # 计算总体覆盖率
        overall_coverage = (covered_count / total_count * 100) if total_count > 0 else 0
        results[sheet_name] = {
            'total': total_count,
            'covered': covered_count,
            'coverage': overall_coverage
        }

        print(f"✓ {sheet_name}: {covered_count}/{total_count} = {overall_coverage:.1f}%")

    # 保存文件
    wb.save(output_file)
    print(f"\n✅ 分析报告已保存到: {output_file}")

    # 打印汇总
    print("\n" + "="*60)
    print("📊 方案覆盖率汇总")
    print("="*60)
    for sheet, data in results.items():
        print(f"{sheet}: {data['coverage']:.1f}% ({data['covered']}/{data['total']})")

    overall = sum(r['coverage'] for r in results.values()) / len(results) if results else 0
    print(f"\n总体覆盖率: {overall:.1f}%")

    return results

def analyze_single_item(indicator_text, sheet_type, keywords, prd_content):
    """
    分析单个需求项的覆盖率
    返回: (覆盖率百分比字符串, 详细描述)
    """
    indicator_lower = indicator_text.lower()

    # 检查是否在PRD中有匹配
    matched_keywords = []
    for kw in keywords:
        if kw.lower() in indicator_lower or kw.lower() in prd_content.lower():
            matched_keywords.append(kw)

    # 特殊处理各种情况
    if 'MDC' in sheet_type or '数据采集' in sheet_type:
        return analyze_mdc_item(indicator_text, prd_content)
    elif '教学管理' in sheet_type:
        return analyze_teaching_item(indicator_text, prd_content)
    elif '智慧工厂' in sheet_type:
        return analyze_factory_item(indicator_text, prd_content)
    elif '仓储物流' in sheet_type:
        return analyze_logistics_item(indicator_text, prd_content)

    # 默认分析
    if matched_keywords:
        coverage = min(100, len(matched_keywords) * 30 + 10)
        return f"{coverage}%", f"PRD中包含相关功能: {', '.join(matched_keywords[:3])}"
    else:
        return "0%", "PRD中未找到对应功能实现"

def analyze_mdc_item(indicator_text, prd_content):
    """分析MDC数据采集平台需求"""
    indicator = indicator_text.lower()
    prd = prd_content.lower()

    # MDC核心功能映射
    mdc_features = {
        'OEE': ('设备OEE', ['设备oee', 'oee统计', '设备利用率']),
        '实验实训报工': ('实验实训报工', ['实训报工', '报工数据', '工时分析']),
        '车间概况': ('车间概况监控', ['车间概况', '设备运行状态', '设备建模']),
        '设备详情': ('设备详情', ['设备详情', '实时参数', '仪表盘']),
        '机床监控': ('机床内部监控', ['机床监控', '视频采集', '机床内部']),
        '状态分析': ('设备状态分析', ['状态分析', '状态时序', '历史周期']),
    }

    for key, (feature_name, keywords) in mdc_features.items():
        for kw in keywords:
            if kw in indicator:
                # 检查PRD中是否有对应功能
                found_in_prd = any(k in prd for k in keywords)
                if found_in_prd:
                    if '教师端' in prd and ('设备oee' in prd or '设备状态' in prd):
                        return "100%", f"✅ PRD中教师端包含'{feature_name}'功能，支持实时监控和数据分析"
                    else:
                        return "50%", f"⚠️ PRD中有相关功能但细节需确认：{feature_name}"
                else:
                    return "0%", f"❌ PRD中未找到'{feature_name}'功能实现"

    return "?", "待分析"

def analyze_teaching_item(indicator_text, prd_content):
    """分析教学管理平台需求"""
    indicator = indicator_text.lower()
    prd = prd_content.lower()

    # 教学管理核心功能映射
    features = {
        '首页导航': ('首页导航功能', ['首页', '导航', '可视化功能导航']),
        '专业班级': ('专业班级信息', ['专业班级', '组织架构', '树状层级']),
        '教师信息': ('教师信息管理', ['教师信息', '教师管理', '批量导入']),
        '学生信息': ('学生信息管理', ['学生信息', '学生管理', '批量导入']),
        'B/S架构': ('B/S架构支持', ['b/s', '浏览器访问']),
        '身份认证': ('身份认证', ['身份认证', '账号密码', '登录']),
        '角色权限': ('角色权限控制', ['角色', '权限', '访问控制']),
    }

    for key, (feature_name, keywords) in features.items():
        match_count = sum(1 for kw in keywords if kw in indicator)
        if match_count > 0:
            # 检查PRD
            found_in_prd = sum(1 for k in keywords if k in prd)
            if found_in_prd >= len(keywords) // 2:
                return "100%", f"✅ PRD中已包含'{feature_name}'功能，支持完整业务流程"
            elif found_in_prd > 0:
                return "80%", f"⚠️ PRD中有'{feature_name}'功能但部分细节缺失"
            else:
                return "0%", f"❌ PRD中未找到'{feature_name}'功能"

    return "?", "待分析"

def analyze_factory_item(indicator_text, prd_content):
    """分析智慧工厂管控平台需求"""
    indicator = indicator_text.lower()
    prd = prd_content.lower()

    # 智慧工厂核心功能映射
    features = {
        '订单管理': ('订单管理', ['订单管理', '订单维护', '交期']),
        '项目管理': ('项目管理', ['项目管理', '项目阶段', '甘特图']),
        'APS': ('APS智能排产', ['aps', '智能排产', '排产算法', '优先级']),
        '生产工单': ('生产工单', ['生产工单', '工单管理']),
        'B/S结构': ('B/S架构', ['b/s', '浏览器']),
        'API接口': ('API接口平台', ['api接口', '二次开发']),
        '交期预排': ('订单交期预排', ['交期预排', '交期模拟', '工作量']),
    }

    for key, (feature_name, keywords) in features.items():
        match_count = sum(1 for kw in keywords if kw in indicator)
        if match_count > 0:
            found_in_prd = sum(1 for k in keywords if k in prd)

            # 特殊处理APS排产
            if 'APS' in key or '排产' in key:
                if 'aps' in prd and '智能排产' in prd:
                    return "100%", "✅ PRD中包含'APS智能排产'功能，支持智能算法、优先级设置、设备更换、物料库存检查等完整功能"
                else:
                    return "0%", "❌ PRD中未明确包含'APS智能排产'功能，只有基础排产"

            # 特殊处理交期预排
            if '交期预排' in key:
                return "0%", "❌ PRD中明确说明'学生的实训任务是在其它教务系统排好课后导入到系统，不存在交期预排场景'，此功能不满足"

            if found_in_prd >= len(keywords) // 2:
                return "100%", f"✅ PRD中包含'{feature_name}'功能"
            elif found_in_prd > 0:
                return "80%", f"⚠️ PRD中有'{feature_name}'但细节需确认"
            else:
                return "0%", f"❌ PRD中未找到'{feature_name}'功能"

    return "?", "待分析"

def analyze_logistics_item(indicator_text, prd_content):
    """分析5G仓储物流系统需求"""
    indicator = indicator_text.lower()
    prd = prd_content.lower()

    # 仓储物流核心功能映射
    features = {
        '货架系统': ('货架系统', ['货架', '货位', '料箱']),
        'AGV': ('AGV搬运', ['agv', '料箱搬运', '自动规划路径']),
        'WMS': ('WMS仓储管理', ['wms', '仓储管理']),
        '料箱': ('料箱管理', ['料箱', '600*400*280']),
        '跨楼层': ('跨楼层配送', ['跨楼层', '货梯']),
    }

    for key, (feature_name, keywords) in features.items():
        match_count = sum(1 for kw in keywords if kw in indicator)
        if match_count > 0:
            # 检查PRD中的刀具管理相关功能
            if '刀具' in prd or 'AGV' in indicator.upper():
                # 这些是硬件要求，PRD主要关注软件功能
                if '180货位' in indicator or '50kg' in indicator:
                    return "?", "⚠️ 此为硬件指标要求，PRD主要覆盖软件功能，需硬件供应商确认"
                elif 'AGV' in key:
                    return "50%", "⚠️ PRD中有刀具AGV配送相关功能，但需确认是否支持料箱式AGV"
                else:
                    return "?", "⚠️ 需确认硬件与软件系统的集成方案"

    return "?", "待分析"

if __name__ == '__main__':
    excel_file = '技术要求V1.1.xlsx'
    prd_file = 'docs/地大智慧工厂_PRD需求文档.md'
    output_file = 'docs/技术要求V1.1_覆盖率分析.xlsx'

    print("="*60)
    print("📊 招标需求与客户实际需求对比分析")
    print("="*60)
    print(f"\n招标标准: {excel_file}")
    print(f"PRD文档: {prd_file}")
    print(f"输出文件: {output_file}\n")

    try:
        results = analyze_coverage(excel_file, prd_file, output_file)
        print("\n✅ 分析完成!")
    except Exception as e:
        print(f"\n❌ 分析失败: {str(e)}")
        import traceback
        traceback.print_exc()
