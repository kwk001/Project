#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从PRD文档提取功能清单并生成Excel文件 - V2改进版
修复功能点名称提取问题
"""

import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# 结构化章节（需要特殊处理，子章节是"功能描述/业务流程"等）
STRUCTURED_SECTIONS = {
    '3.2', '3.3', '3.7', '3.8', '3.12', '3.13', '3.14', '3.15',
    '3.17', '3.18', '3.19', '3.20', '3.21', '3.22', '3.23', '3.24', '3.25', '3.26'
}

# 章节到功能点的映射（手动维护关键章节的详细功能）
MODULE_FUNCTIONS = {
    '3.2 智能卡口控制': [
        {'name': '车牌自动识别', 'desc': '摄像头自动识别车牌号，识别率≥95%，支持人工补录'},
        {'name': '车辆自动称重', 'desc': '地磅自动采集毛重，与预报重量差异超±5%触发预警'},
        {'name': '海关验核对碰', 'desc': '向金关二期发送验核请求，接收布控/放行/禁止指令'},
        {'name': '自动道闸控制', 'desc': '根据验核结果自动控制道闸开启/关闭，记录通行日志'},
        {'name': '布控车辆拦截', 'desc': '布控车辆自动拦截，下发查验指令，推送至查验区域'},
        {'name': '通行记录查询', 'desc': '查询车辆通行记录，支持按时间/车牌/结果筛选'},
        {'name': '卡口设备监控', 'desc': '监控摄像头/地磅/道闸状态，设备故障自动告警'},
    ],
    '3.3 查验预警处置': [
        {'name': '查验预警提示', 'desc': '货物到港前自动提示查验要求，引导车辆至查验区'},
        {'name': '查验任务分配', 'desc': '自动分配查验任务至查验员，支持人工调整'},
        {'name': '查验作业记录', 'desc': '记录查验过程，拍照留档，录入查验结果'},
        {'name': '查验异常处置', 'desc': '异常货物处理流程，支持退运/销毁/扣留等操作'},
        {'name': '查验结果反馈', 'desc': '查验结果实时反馈海关，触发放行或扣留指令'},
        {'name': '查验直放处理', 'desc': '查验正常货物快速放行，减少二次搬运'},
    ],
    '3.5 监管仓管理': [
        {'name': '收货记录管理', 'desc': '记录收货信息，关联预报单，生成入库任务'},
        {'name': '上架单管理', 'desc': '生成上架任务，分配库位，指导上架作业'},
        {'name': '库存实时查询', 'desc': '实时查询库存数量、位置、状态，支持多维度筛选'},
        {'name': '库存盘点管理', 'desc': '定期/不定期盘点，盘点差异处理，盈亏调整'},
        {'name': '库存移位管理', 'desc': '库位间货物移位，移位单管理，位置变更记录'},
        {'name': '发货单管理', 'desc': '创建发货单，分配出库任务，指导拣货作业'},
        {'name': '打板作业管理', 'desc': '航空打板作业指导，板号管理，重量体积校验'},
        {'name': '上机确认管理', 'desc': '货物装机确认，航班配载，装机单生成'},
        {'name': '海关查验协同', 'desc': '查验货物定位，查验结果同步，放行后重新入库'},
    ],
    '3.7 订单管理': [
        {'name': '订单创建与审核', 'desc': '客户下单，订单信息审核，审核通过生成运单'},
        {'name': '订单变更管理', 'desc': '支持收货信息、装卸货日期等变更，接口下发变更'},
        {'name': '订单拆分管理', 'desc': '按明细/定量模式拆分主订单，支持取消拆分'},
        {'name': '订单可视化跟踪', 'desc': '订单全流程跟踪，集货时间监控，协同看板展示'},
    ],
    '3.8 运单管理': [
        {'name': '运单创建与变更', 'desc': '根据订单生成运单，支持运单信息变更'},
        {'name': '运单跟踪查询', 'desc': '运单全流程跟踪，在途位置监控，节点状态更新'},
        {'name': '运单自动化配置', 'desc': '自动化规则配置，自动分配承运商/车辆/司机'},
        {'name': '运单拆分与配载', 'desc': '运段拆分，整车/零担配载，运力优化'},
    ],
    '3.12 运行中心-综合态势大屏': [
        {'name': '综合态势展示', 'desc': '业务/交通/安防/能耗/安全/消防/资产/设备态势综合展示'},
        {'name': '交通态势监控', 'desc': '车辆进出流量、车位占用、排队等待时间实时监控'},
        {'name': '作业态势监控', 'desc': '仓库作业量、月台利用率、装卸效率实时监控'},
        {'name': '异常预警展示', 'desc': '各类异常事件汇总展示，红色预警优先显示'},
    ],
    '3.13 客服业务管理': [
        {'name': '在线客服咨询', 'desc': '客户在线咨询，即时消息回复，问题解答'},
        {'name': '工单管理', 'desc': '异常问题登记，工单分派，处理跟踪，结果回访'},
        {'name': '客户反馈管理', 'desc': '客户投诉建议收集，处理进度跟踪，满意度评价'},
        {'name': '知识库管理', 'desc': '常见问题库，自助查询，快速回复模板'},
    ],
    '3.14 登录功能': [
        {'name': '账号密码登录', 'desc': '支持用户名/密码登录，密码加密存储'},
        {'name': '短信验证码登录', 'desc': '手机号+短信验证码登录'},
        {'name': '单点登录集成', 'desc': 'CAS/SSO单点登录，统一身份认证'},
    ],
    '3.15 合同管理': [
        {'name': '合同模板管理', 'desc': '合同模板创建、编辑、版本管理'},
        {'name': '合同签订', 'desc': '电子签章，在线签署，合同生效'},
        {'name': '合同变更', 'desc': '合同条款变更，变更审批，变更记录'},
    ],
    '3.17 接口对接管理': [
        {'name': '金关二期接口', 'desc': '海关申报、验核、布控、放行数据交换'},
        {'name': '安检系统接口', 'desc': '安检申报、结果接收、异常处理'},
        {'name': '航空公司接口', 'desc': '航班动态、舱单数据、货物状态对接'},
        {'name': '场站系统接口', 'desc': '货物进出港、仓储状态、航班信息同步'},
        {'name': '公路/铁路平台接口', 'desc': '多式联运数据交换，运单跟踪对接'},
    ],
    '3.18 分拣线控制管理': [
        {'name': '分拣任务下发', 'desc': 'WMS向PLC下发分拣任务，货物路由规则'},
        {'name': '扫码识别', 'desc': '货物条码扫描，自动识别货物信息'},
        {'name': '自动分拣控制', 'desc': 'PLC控制分拣设备，货物分配至指定格口'},
        {'name': '查验货物分流', 'desc': '布控货物自动分流至查验区'},
        {'name': '打板任务管理', 'desc': '打板任务分配，打板进度跟踪'},
        {'name': '分拣数据监控', 'desc': '分拣效率统计，设备状态监控，异常告警'},
        {'name': '海关数据对接', 'desc': '分拣数据实时推送海关，查验结果同步'},
    ],
    '3.19 场站系统对接': [
        {'name': '航班信息同步', 'desc': '航班计划、动态信息实时同步'},
        {'name': '舱单数据交换', 'desc': '进出口舱单数据对接，申报状态同步'},
        {'name': '货物状态同步', 'desc': '货物收运、存储、装机状态实时同步'},
        {'name': '费用数据对接', 'desc': '货站费用计算，场站结算数据交换'},
    ],
    '3.20 航空物流公共信息平台': [
        {'name': '出港服务', 'desc': '货物出港信息查询，在线订舱，运价查询'},
        {'name': '进港服务', 'desc': '货物进港查询，提货预约，费用缴纳'},
        {'name': '全程可视化', 'desc': '货物运输全流程跟踪，节点状态推送'},
        {'name': '通关物流协同', 'desc': '通关状态查询，物流状态联动，一站式服务平台'},
        {'name': '账号权限管理', 'desc': '企业/个人账号注册，权限划分，安全认证'},
    ],
    '3.21 移动端功能': [
        {'name': '车辆预约', 'desc': '司机APP预约进港时间，预约单管理'},
        {'name': '签到排队', 'desc': '到场扫码签到，排队叫号通知'},
        {'name': '作业任务', 'desc': '接收作业任务，完成任务反馈，异常情况上报'},
        {'name': '消息通知', 'desc': '预约提醒、叫号通知、异常提醒推送'},
    ],
    '3.22 统一门户': [
        {'name': '应用集成', 'desc': '各业务系统统一入口，单点登录'},
        {'name': '待办集成', 'desc': '各系统待办任务汇总，统一处理'},
        {'name': '消息集成', 'desc': '系统消息统一展示，消息推送'},
    ],
    '3.23 运维服务管理': [
        {'name': 'IT运维工单', 'desc': '系统故障报修，工单分派，处理跟踪'},
        {'name': '设施运维工单', 'desc': '设施设备报修，维修调度，验收确认'},
        {'name': '巡检计划', 'desc': '定期巡检计划，巡检任务执行，巡检记录'},
    ],
    '3.24 系统管理': [
        {'name': '用户管理', 'desc': '用户增删改查，状态管理，密码策略'},
        {'name': '角色权限管理', 'desc': '角色定义，权限分配，数据权限控制'},
        {'name': '操作日志', 'desc': '操作日志记录，日志查询，日志导出'},
        {'name': '登录日志', 'desc': '登录记录，异常登录告警'},
    ],
    '3.25 跨境电商业务管理': [
        {'name': '9610三单对碰', 'desc': '订单/支付单/物流单三单校验，数据匹配'},
        {'name': '9610清单申报', 'desc': '清单数据生成，清单申报，汇总申报'},
        {'name': '9710订单备案', 'desc': 'B2B订单备案，报关单生成，退税申报'},
        {'name': '9810海外仓备案', 'desc': '海外仓企业备案，备货申报，退货管理'},
        {'name': '跨境电商分拣', 'desc': '跨境电商货物特殊分拣规则，查验分流'},
        {'name': '海关数据对接', 'desc': '跨境电商通关服务平台数据交换'},
    ],
    '3.26 区区联动管理': [
        {'name': '转关申请管理', 'desc': '转关单申请，海关审批，施封管理'},
        {'name': '途中监控', 'desc': 'GPS位置跟踪，路线偏离预警，超时告警'},
        {'name': '运抵确认', 'desc': '联运区运抵确认，自动解封，入库处理'},
        {'name': '综保区账册', 'desc': '保税货物账册管理，入区/出区申报'},
        {'name': '联动数据展示', 'desc': '区区联动数据大屏，监控指标展示'},
    ],
}


def extract_prd_content(file_path):
    """读取PRD文档内容"""
    with open(file_path, 'r', encoding='utf-8') as f:
        return f.read()


def extract_chapters(content):
    """提取章节结构"""
    chapters = []
    pattern = r'^### (3\.\d+(?:\.\d+)?)\s+(.+?)$'
    matches = list(re.finditer(pattern, content, re.MULTILINE))

    for i, match in enumerate(matches):
        chapter_num = match.group(1)
        chapter_title = match.group(2).replace('⭐新增', '').replace('⭐完善', '').replace('（v2.0）', '').replace('⭐新增模块', '').strip()
        start = match.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(content)

        chapters.append({
            'num': chapter_num,
            'title': chapter_title,
            'full_title': f"{chapter_num} {chapter_title}",
            'start': start,
            'end': end,
            'content': content[start:end]
        })

    return chapters


def extract_subsections(chapter_content):
    """提取子章节（#### 级别的内容）"""
    subsections = []
    pattern = r'^#### (\d+\.\d+\.\d+)\s+(.+?)$'
    matches = list(re.finditer(pattern, chapter_content, re.MULTILINE))

    for i, match in enumerate(matches):
        sub_num = match.group(1)
        sub_title = match.group(2).replace('⭐新增', '').replace('⭐完善', '').replace('（v2.0）', '').replace('⭐新增模块', '').replace('⭐补充字段', '').strip()
        start = match.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(chapter_content)

        subsections.append({
            'num': sub_num,
            'title': sub_title,
            'content': chapter_content[start:end]
        })

    return subsections


def get_priority(module_name):
    """根据模块名称判断优先级"""
    if any(x in module_name for x in ['基础资料', '智能卡口', '查验预警', '货物预报', '监管仓', '跨境电商']):
        return 'P0'
    elif any(x in module_name for x in ['新增', '完善']):
        return 'P1'
    else:
        return 'P1'


def get_complexity(func_name, module_name, is_new=False):
    """判断复杂度"""
    if '对接' in func_name or '接口' in func_name or 'PLC' in func_name or '控制' in func_name:
        return '高'
    elif is_new:
        return '中'
    elif any(x in func_name for x in ['管理', '配置', '同步']):
        return '中'
    else:
        return '低'


def generate_function_list(prd_file):
    """生成功能清单数据"""
    content = extract_prd_content(prd_file)
    chapters = extract_chapters(content)

    function_list = []
    func_id = 1

    for ch in chapters:
        # 跳过三级章节（如3.9.5是3.9的子章节）
        if ch['num'].count('.') == 2:
            continue

        module_key = ch['full_title']

        # 检查是否有预定义的功能列表
        if module_key in MODULE_FUNCTIONS:
            for func in MODULE_FUNCTIONS[module_key]:
                function_list.append({
                    'ID': f'F{func_id:03d}',
                    '一级目录': ch['title'],
                    '二级目录': '',
                    '页面名称': '',
                    '功能点': func['name'],
                    '详细描述': func['desc'],
                    '优先级': get_priority(ch['title']),
                    '复杂度': get_complexity(func['name'], ch['title']),
                    '依赖': '',
                    '验收标准': ''
                })
                func_id += 1
        else:
            # 提取子章节作为功能点
            subsections = extract_subsections(ch['content'])

            if subsections:
                for sub in subsections:
                    # 跳过结构化章节的非功能子章节
                    if ch['num'] in STRUCTURED_SECTIONS and sub['title'] in ['功能描述', '业务流程', '业务规则', '验收标准', '数据库设计', '接口设计']:
                        continue

                    # 提取功能描述
                    desc_match = re.search(r'\*\*功能描述\*\*[\s\S]*?-\s+\*\*功能价值\*\*：(.+?)(?:\n|$)', sub['content'])
                    detail = desc_match.group(1).strip() if desc_match else ''

                    # 如果没有功能描述段落，提取第一段内容
                    if not detail:
                        para_match = re.search(r'^####[\s\S]*?\n\n(.+?)(?:\n\n|\n###|\Z)', sub['content'], re.DOTALL)
                        if para_match:
                            detail = para_match.group(1).strip()[:200]

                    function_list.append({
                        'ID': f'F{func_id:03d}',
                        '一级目录': ch['title'],
                        '二级目录': sub['title'],
                        '页面名称': '',
                        '功能点': sub['title'],
                        '详细描述': detail,
                        '优先级': get_priority(ch['title']),
                        '复杂度': get_complexity(sub['title'], ch['title']),
                        '依赖': '',
                        '验收标准': ''
                    })
                    func_id += 1
            else:
                # 没有子章节，提取章节本身的功能描述
                desc_match = re.search(r'\*\*功能描述\*\*[\s\S]*?-\s+\*\*功能价值\*\*：(.+?)(?:\n|$)', ch['content'])
                detail = desc_match.group(1).strip() if desc_match else ''

                function_list.append({
                    'ID': f'F{func_id:03d}',
                    '一级目录': ch['title'],
                    '二级目录': '',
                    '页面名称': '',
                    '功能点': ch['title'],
                    '详细描述': detail,
                    '优先级': get_priority(ch['title']),
                    '复杂度': get_complexity(ch['title'], ch['title']),
                    '依赖': '',
                    '验收标准': ''
                })
                func_id += 1

    return function_list


def create_excel(function_list, output_file):
    """创建Excel文件"""
    df = pd.DataFrame(function_list)

    # 创建Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='功能清单')

        # 获取工作表
        workbook = writer.book
        worksheet = writer.sheets['功能清单']

        # 设置表头样式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # 设置列宽
        column_widths = {
            'A': 8,   # ID
            'B': 22,  # 一级目录
            'C': 22,  # 二级目录
            'D': 15,  # 页面名称
            'E': 28,  # 功能点
            'F': 55,  # 详细描述
            'G': 8,   # 优先级
            'H': 8,   # 复杂度
            'I': 12,  # 依赖
            'J': 35   # 验收标准
        }

        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        # 设置数据行样式
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)

        # 冻结首行
        worksheet.freeze_panes = 'A2'

    print(f"功能清单已生成: {output_file}")
    print(f"共 {len(function_list)} 个功能点")


if __name__ == '__main__':
    prd_file = '/Users/apple/Desktop/文件/kaiwu-richang/docs/PRD文档_v2.0.md'
    output_file = '/Users/apple/Desktop/文件/kaiwu-richang/docs/功能清单_v2.0.xlsx'

    function_list = generate_function_list(prd_file)
    create_excel(function_list, output_file)

    # 打印统计
    print("\n功能清单统计：")
    print("=" * 60)

    # 按一级目录统计
    from collections import Counter
    module_counts = Counter([f['一级目录'] for f in function_list])
    for module, count in module_counts.most_common():
        print(f"{module:25} {count:3}个功能点")

    print("=" * 60)

    # 打印前20个功能点预览
    print("\n功能清单预览（前20条）：")
    print("=" * 100)
    for func in function_list[:20]:
        print(f"{func['ID']} | {func['一级目录'][:15]:15} | {func['功能点'][:30]}")
