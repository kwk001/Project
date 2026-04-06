#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从PRD文档提取功能清单并生成Excel文件
"""

import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def extract_prd_content(file_path):
    """读取PRD文档内容"""
    with open(file_path, 'r', encoding='utf-8') as f:
        return f.read()

def extract_chapters(content):
    """提取章节结构"""
    chapters = []
    pattern = r'^### (3\.\d+(?:\.\d+)?)\s+(.+?)$'
    matches = list(re.finditer(pattern, content, re.MULTILINE))

    for i, match in enumerate(matches):
        chapter_num = match.group(1)
        chapter_title = match.group(2).replace('⭐新增', '').replace('⭐完善', '').replace('（v2.0）', '').replace('⭐新增模块', '').strip()
        start = match.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(content)

        chapters.append({
            'num': chapter_num,
            'title': chapter_title,
            'start': start,
            'end': end,
            'content': content[start:end]
        })

    return chapters

def extract_subsections(chapter_content):
    """提取子章节（#### 级别的内容）"""
    subsections = []
    pattern = r'^#### (\d+\.\d+\.\d+)\s+(.+?)$'
    matches = list(re.finditer(pattern, chapter_content, re.MULTILINE))

    for i, match in enumerate(matches):
        sub_num = match.group(1)
        sub_title = match.group(2)
        start = match.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(chapter_content)

        subsections.append({
            'num': sub_num,
            'title': sub_title,
            'content': chapter_content[start:end]
        })

    return subsections

def extract_features_from_chapter(chapter):
    """从章节中提取功能点"""
    features = []

    # 提取子章节作为功能点
    subsections = extract_subsections(chapter['content'])

    if subsections:
        for sub in subsections:
            # 提取功能描述
            desc_match = re.search(r'\*\*功能描述\*\*[\s\S]*?-\s+\*\*功能名称\*\*：(.+?)(?:\n|$)', sub['content'])
            func_name = desc_match.group(1).strip() if desc_match else sub['title']

            # 提取详细描述
            detail_match = re.search(r'\*\*功能描述\*\*[\s\S]*?-\s+\*\*功能价值\*\*：(.+?)(?:\n|$)', sub['content'])
            detail = detail_match.group(1).strip() if detail_match else ''

            # 如果没有功能描述段落，提取段落内容
            if not detail:
                para_match = re.search(r'^####[\s\S]*?\n\n(.+?)(?:\n\n|\n###|\Z)', sub['content'], re.DOTALL)
                if para_match:
                    detail = para_match.group(1).strip()[:200]

            features.append({
                'subsection': sub['title'],
                'func_name': func_name,
                'detail': detail
            })
    else:
        # 没有子章节，将整个章节作为一个功能点
        # 尝试提取功能描述
        desc_match = re.search(r'\*\*功能描述\*\*[\s\S]*?-\s+\*\*功能名称\*\*：(.+?)(?:\n|$)', chapter['content'])
        func_name = desc_match.group(1).strip() if desc_match else chapter['title']

        detail_match = re.search(r'\*\*功能描述\*\*[\s\S]*?-\s+\*\*功能价值\*\*：(.+?)(?:\n|$)', chapter['content'])
        detail = detail_match.group(1).strip() if detail_match else ''

        features.append({
            'subsection': '',
            'func_name': func_name,
            'detail': detail
        })

    return features

def generate_function_list(prd_file):
    """生成功能清单数据"""
    content = extract_prd_content(prd_file)
    chapters = extract_chapters(content)

    function_list = []
    func_id = 1

    for ch in chapters:
        # 跳过非主要章节（如3.9.5是3.9的子章节）
        if '.' in ch['num'] and ch['num'].count('.') == 2:
            continue

        # 提取功能点
        features = extract_features_from_chapter(ch)

        for feat in features:
            function_list.append({
                'ID': f'F{func_id:03d}',
                '一级目录': ch['title'],
                '二级目录': feat['subsection'] if feat['subsection'] else '',
                '页面名称': '',
                '功能点': feat['func_name'],
                '详细描述': feat['detail'],
                '优先级': 'P0' if '新增' in ch['title'] or ch['num'] in ['3.1', '3.2', '3.3', '3.4', '3.5'] else 'P1',
                '复杂度': '高' if '新增' in ch['title'] else '中',
                '依赖': '',
                '验收标准': ''
            })
            func_id += 1

    return function_list

def create_excel(function_list, output_file):
    """创建Excel文件"""
    df = pd.DataFrame(function_list)

    # 创建Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='功能清单')

        # 获取工作表
        workbook = writer.book
        worksheet = writer.sheets['功能清单']

        # 设置表头样式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # 设置列宽
        column_widths = {
            'A': 8,   # ID
            'B': 20,  # 一级目录
            'C': 25,  # 二级目录
            'D': 15,  # 页面名称
            'E': 25,  # 功能点
            'F': 50,  # 详细描述
            'G': 8,   # 优先级
            'H': 8,   # 复杂度
            'I': 10,  # 依赖
            'J': 30   # 验收标准
        }

        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        # 设置数据行样式
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)

        # 冻结首行
        worksheet.freeze_panes = 'A2'

    print(f"功能清单已生成: {output_file}")
    print(f"共 {len(function_list)} 个功能点")

if __name__ == '__main__':
    prd_file = '/Users/apple/Desktop/文件/kaiwu-richang/docs/PRD文档_v2.0.md'
    output_file = '/Users/apple/Desktop/文件/kaiwu-richang/docs/功能清单_v2.0.xlsx'

    function_list = generate_function_list(prd_file)
    create_excel(function_list, output_file)

    # 打印前10个功能点预览
    print("\n功能清单预览（前10条）：")
    print("=" * 100)
    for func in function_list[:10]:
        print(f"{func['ID']} | {func['一级目录']} | {func['功能点']}")
